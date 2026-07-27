"""Экспорт результатов keyword research в .xlsx (openpyxl). READ-ONLY, без секретов.

Один лист «Ключевые слова»: строки сгруппированы по кластеру (интент), внутри — по убыванию
объёма поиска. Метрики уже посчитаны кодом (ads.keyword_plan.KeywordIdea); micros→валюту тоже
считал КОД. Стиль шапки/форматов — как в reports.xlsx (единый вид экспорта).
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ads.keyword_plan import KeywordIdea, seasonality_sparkline
from confirm.attachment import MAX_ATTACHMENT_ROWS  # потолок строк — политика, не эта функция
from core.logging import redact_text  # правило 5: .xlsx уезжает в Telegram навсегда
from keywords.cluster import Cluster
from reports.safe_cell import safe_row  # A9: защита от формула-инъекции в ячейках


def _append(ws, row) -> None:
    """A9: ws.append с обезвреживанием ячеек (safe_row) — имена/ключи из данных клиента."""
    ws.append(safe_row(row))


def _writerow(w, row) -> None:
    """A9: csv.writerow с обезвреживанием ячеек (safe_row) — CSV injection в тексте ключа."""
    w.writerow(safe_row(row))


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _headers(currency: str = "", with_relevance: bool = False) -> list[str]:
    """D8: заголовки таблицы. Денежные колонки несут ВАЛЮТУ АККАУНТА (ставки в micros→валюту уже
    посчитал КОД, ads.keyword_plan) — раньше «Ставка/CPC» без валюты вводили в заблуждение при
    не-USD аккаунте. Пустая валюта (тест-аккаунт/сбой чтения) → без суффикса, как раньше.
    with_relevance (§19.4.2) добавляет колонку «Релевантность» В КОНЕЦ (позиция 12) — чтобы НЕ
    сдвигать индексы _NUM_FORMATS/спарклайна; включается только когда есть вердикты."""
    suf = f" ({currency})" if currency else ""
    cols = [
        "Кластер",
        "Интент",
        "Ключевое слово",
        "Объём/мес",
        "Конкуренция",
        "Индекс (0-100)",
        f"Сред. CPC{suf}",
        f"Ставка низ{suf}",
        f"Ставка верх{suf}",
        "Пик сезона",
        "Сезонность (12 мес)",  # §7: полная кривая как unicode-спарклайн
    ]
    # §19.4.2: колонка «Релевантность» (релевантно | низкая) — только когда есть вердикты.
    return cols + (["Релевантность"] if with_relevance else [])


HEADERS = _headers()  # обратная совместимость (тесты/старые вызовы без валюты)
# Форматы числовых колонок (с позиции «Объём/мес»); «Пик сезона» (10)/«Сезонность» (11)/
# «Релевантность» (12) — текст.
_NUM_FORMATS = {4: "#,##0", 6: "0", 7: "#,##0.00", 8: "#,##0.00", 9: "#,##0.00"}


def _relevance_cell(relevance: dict | None, text: str) -> str:
    """§19.4.2: текст ячейки релевантности. Fail-open: отсутствующее/пустое → «релевантно»."""
    return "релевантно" if (relevance or {}).get(text, True) else "низкая"


def _autosize(ws, ncols: int, *, cap: int = 50) -> None:
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        width = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), cap)


def build_workbook(
    clusters: list[Cluster],
    ideas: list[KeywordIdea],
    *,
    seeds: list[str] | None = None,
    url: str | None = None,
    language: str = "ru",
    negatives: list[str] | None = None,
    currency: str = "",
    relevance: dict | None = None,
) -> Workbook:
    by_text = {i.text: i for i in ideas}  # метрики по тексту ключа
    with_rel = relevance is not None  # §19.4.2: колонка вердикта только когда есть вердикты
    headers = _headers(currency, with_rel)  # D8: денежные колонки с валютой аккаунта
    wb = Workbook()
    ws = wb.active
    ws.title = "Ключевые слова"

    src = ", ".join(seeds or []) or (url or "")
    _append(ws, [f"Keyword research — сиды: {src} (язык: {language})"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _append(ws, [f"Всего ключей: {len(by_text)}, кластеров: {len(clusters)}"])
    _append(ws, [])

    header_row = ws.max_row + 1
    _append(ws, headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).fill = _HEADER_FILL
        ws.cell(row=header_row, column=c).font = _HEADER_FONT
    ws.freeze_panes = f"A{header_row + 1}"

    for cl in clusters:
        rows = sorted(
            (by_text.get(k) for k in cl.keywords),
            key=lambda i: (i.avg_monthly_searches if i else 0),
            reverse=True,
        )
        for idea in rows:
            if idea is None:
                continue
            row = [
                cl.name,
                cl.intent,
                idea.text,
                idea.avg_monthly_searches,
                idea.competition,
                idea.competition_index,
                round(idea.avg_cpc, 2),
                round(idea.low_bid, 2),
                round(idea.high_bid, 2),
                idea.peak_month,
                seasonality_sparkline(getattr(idea, "monthly", None)),
            ] + ([_relevance_cell(relevance, idea.text)] if with_rel else [])
            _append(ws, row)

    for col, fmt in _NUM_FORMATS.items():
        for r in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=r, column=col).number_format = fmt
    _autosize(ws, len(headers))

    # §7 «предложение минус-слов» — отдельный лист (advisory; добавление идёт командой за confirm-гейтом).
    if negatives:
        ws2 = wb.create_sheet("Минус-слова")
        _append(
            ws2, ["Предложенные минус-слова (добавляются отдельной командой после подтверждения)"]
        )
        ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
        _append(ws2, [])
        hr = ws2.max_row + 1
        _append(ws2, ["#", "Минус-слово"])
        for c in range(1, 3):
            ws2.cell(row=hr, column=c).fill = _HEADER_FILL
            ws2.cell(row=hr, column=c).font = _HEADER_FONT
        ws2.freeze_panes = f"A{hr + 1}"
        for i, n in enumerate(negatives, 1):
            _append(ws2, [i, n])
        _autosize(ws2, 2)
    return wb


def write_keywords_xlsx(
    clusters: list[Cluster],
    ideas: list[KeywordIdea],
    path: str,
    *,
    seeds: list[str] | None = None,
    url: str | None = None,
    language: str = "ru",
    negatives: list[str] | None = None,
    currency: str = "",
    relevance: dict | None = None,
) -> str:
    """Сохранить .xlsx по пути path. Возвращает path. currency (D8) — валюта аккаунта в колонках ставок.
    relevance (§19.4.2) — {ключ: релевантно?}; задан ⇒ добавляется колонка «Релевантность»."""
    build_workbook(
        clusters,
        ideas,
        seeds=seeds,
        url=url,
        language=language,
        negatives=negatives,
        currency=currency,
        relevance=relevance,
    ).save(path)
    return path


def write_keywords_csv(
    clusters: list[Cluster],
    ideas: list[KeywordIdea],
    path: str,
    *,
    seeds: list[str] | None = None,
    url: str | None = None,
    language: str = "ru",
    currency: str = "",
    relevance: dict | None = None,
) -> str:
    """ТЗ §7 «CSV/table export»: те же данные, что .xlsx, но плоским CSV (для импорта в таблицы/BI).
    Кодировка utf-8-sig (BOM) — Excel корректно открывает кириллицу. Группировка по кластеру, внутри —
    по убыванию объёма (как в .xlsx). Возвращает path. Минус-слова сюда не кладём (advisory, отдельно).
    relevance (§19.4.2) — {ключ: релевантно?}; задан ⇒ добавляется колонка «Релевантность»."""
    import csv

    by_text = {i.text: i for i in ideas}
    with_rel = relevance is not None
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        _writerow(w, _headers(currency, with_rel))  # D8: денежные колонки с валютой аккаунта
        for cl in clusters:
            rows = sorted(
                (by_text.get(k) for k in cl.keywords),
                key=lambda i: (i.avg_monthly_searches if i else 0),
                reverse=True,
            )
            for idea in rows:
                if idea is None:
                    continue
                row = [
                    cl.name,
                    cl.intent,
                    idea.text,
                    idea.avg_monthly_searches,
                    idea.competition,
                    idea.competition_index,
                    round(idea.avg_cpc, 2),
                    round(idea.low_bid, 2),
                    round(idea.high_bid, 2),
                    idea.peak_month,
                    seasonality_sparkline(getattr(idea, "monthly", None)),
                ] + ([_relevance_cell(relevance, idea.text)] if with_rel else [])
                _writerow(w, row)
    return path


def build_keyword_list_workbook(
    keywords: list[str],
    match_type: str,
    action: str,
    *,
    scope: str = "",
    max_rows: int = MAX_ATTACHMENT_ROWS,
) -> Workbook:
    """Компактный список ключей/минус-слов для ЧЕРНОВИКА мутации (ТЗ §5: большие списки —
    вложением, не текстом). Одна вкладка: # / ключ / тип соответствия / действие (+ «Кампания /
    Общий набор», если известна). Без метрик и секретов. Возвращает Workbook — НЕ путь.

    Сплит «собрать» и «сохранить» не косметический: собирает файл процесс-курьер (`scheduler`), а
    решение приложить его принимает тул-слой в другом контейнере — путь на ФС границу процессов не
    переживает. Плюс L3-график (Волна 5) цепляется к `Workbook`, а к пути — не к чему.

    `max_rows` — потолок СТРОК, а не байтов: список приходит из текста менеджера/модели, и без
    потолка вставленные из CSV 200k минус-слов дают файл больше лимита Telegram 50 МБ, то есть
    обещание без файла. Усечение не молчит — в лист уходит явная пометка.

    Ключ проходит `redact_text`: текст приходит от человека/модели, а .xlsx уезжает в Telegram
    навсегда и отзыву не подлежит (правило 5). `safe_row` рядом занимается формула-инъекцией, это
    другой класс и он не заменяет редакцию."""
    total = len(keywords)
    rows = list(keywords)[: max(0, int(max_rows))]
    with_scope = bool(scope)
    wb = Workbook()
    ws = wb.active
    ws.title = "Черновик ключей"
    head = f"{action} — {total} шт. (тип соответствия: {match_type})"
    _append(ws, [redact_text(head + (f" · {scope}" if with_scope else ""))])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    if len(rows) < total:
        _append(ws, [f"⚠️ показаны первые {len(rows)} из {total} — список усечён"])
    _append(ws, [])

    header_row = ws.max_row + 1
    # Конкатенация, а не `.append`: гард формула-инъекции (tests/test_export_formula_injection.py)
    # запрещает в этом файле любой `.append(` мимо `safe_row` — намеренно грубо, разбором ТЕКСТА, без
    # разбора типа. Обходить гард ради косметики списка неверно: он ловит именно тот класс, где
    # «это же не лист, а обычный list» оказывается ошибкой.
    headers = ["#", "Ключевое слово", "Тип соответствия", "Действие"] + (
        ["Кампания / Общий набор"] if with_scope else []
    )
    _append(ws, headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).fill = _HEADER_FILL
        ws.cell(row=header_row, column=c).font = _HEADER_FONT
    ws.freeze_panes = f"A{header_row + 1}"

    for i, kw in enumerate(rows, 1):
        row = [i, redact_text(str(kw)), match_type, action] + ([scope] if with_scope else [])
        _append(ws, row)
    _autosize(ws, len(headers))
    return wb


def write_keyword_list_xlsx(
    keywords: list[str], match_type: str, action: str, path: str, *, scope: str = ""
) -> str:
    """Сохранить список ключей черновика по пути path. Возвращает path. Тонкая обёртка над
    `build_keyword_list_workbook` — сигнатура обратно совместима (`scope` kw-only с дефолтом)."""
    build_keyword_list_workbook(keywords, match_type, action, scope=scope).save(path)
    return path
