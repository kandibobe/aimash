from __future__ import annotations

import asyncio
from datetime import date
from types import SimpleNamespace

from openpyxl import load_workbook

from ads.client import DRAFT_ACCOUNT_ID
from reports.keyword_export import (
    KeywordPerformanceRow,
    fetch_keyword_performance_rows,
    write_keyword_performance_xlsx,
)
from reports.period import Period


class _Service:
    def __init__(self, rows):
        self.rows = rows
        self.calls = []

    def search_stream(self, *, customer_id, query):
        self.calls.append((customer_id, query))
        return [SimpleNamespace(results=self.rows)]


class _Client:
    def __init__(self, service):
        self.service = service

    def get_service(self, name):
        assert name == "GoogleAdsService"
        return self.service


def _sdk_row(keyword: str, *, cost_micros: int = 1_500_000):
    return SimpleNamespace(
        campaign=SimpleNamespace(name="Search campaign"),
        ad_group=SimpleNamespace(name="Group A"),
        ad_group_criterion=SimpleNamespace(
            keyword=SimpleNamespace(text=keyword, match_type=SimpleNamespace(name="PHRASE"))
        ),
        metrics=SimpleNamespace(
            impressions=100,
            clicks=7,
            cost_micros=cost_micros,
            conversions=2.5,
        ),
    )


def test_keyword_reader_uses_uncapped_keyword_view_search_stream(monkeypatch):
    monkeypatch.setattr("reports.keyword_export.ensure_read_allowed", lambda _account: None)
    service = _Service([_sdk_row("alpha"), _sdk_row("beta")])
    period = Period(date(2025, 1, 1), date(2025, 12, 31), "2025")

    rows = fetch_keyword_performance_rows(_Client(service), DRAFT_ACCOUNT_ID, period)

    assert [row.keyword_text for row in rows] == ["alpha", "beta"]
    assert len(service.calls) == 1
    customer_id, query = service.calls[0]
    assert customer_id == DRAFT_ACCOUNT_ID
    for field in (
        "campaign.name",
        "ad_group.name",
        "ad_group_criterion.keyword.text",
        "ad_group_criterion.keyword.match_type",
        "metrics.impressions",
        "metrics.clicks",
        "metrics.cost_micros",
        "metrics.conversions",
    ):
        assert field in query
    assert "FROM keyword_view" in query
    assert "campaign.advertising_channel_type = 'SEARCH'" in query
    assert "segments.date BETWEEN '2025-01-01' AND '2025-12-31'" in query
    assert "LIMIT" not in query.upper()


def test_keyword_xlsx_writes_one_data_row_per_keyword(tmp_path):
    rows = [
        KeywordPerformanceRow("Campaign", "Group", "alpha", "EXACT", 10, 2, 1_500_000, 1.0),
        KeywordPerformanceRow("Campaign", "Group", "beta", "BROAD", 20, 3, 2_750_000, 0.0),
        KeywordPerformanceRow("Campaign", "Other", "gamma", "PHRASE", 30, 4, 3_000_000, 2.0),
    ]
    path = tmp_path / "keywords.xlsx"

    write_keyword_performance_xlsx(rows, str(path), currency="UAH")
    ws = load_workbook(path)["Ключевые слова"]

    assert ws.max_row == len(rows) + 1
    assert [ws.cell(row=row, column=3).value for row in range(2, 5)] == [
        "alpha",
        "beta",
        "gamma",
    ]
    assert ws.cell(row=2, column=7).value == 1.5


def test_export_keyword_tool_builds_dedicated_row_artifact(monkeypatch, tmp_path):
    from mcp_server import tools_workflows as tw

    monkeypatch.setattr(tw, "ensure_read_allowed", lambda _account: None)
    monkeypatch.setattr(tw, "build_client_async", lambda _account: _async_value(object()))
    period = Period(date(2025, 1, 1), date(2025, 12, 31), "2025")
    monkeypatch.setattr(tw, "account_period", lambda *_args, **_kwargs: _async_value(period))
    rows = [KeywordPerformanceRow("C", "G", "alpha", "EXACT", 1, 1, 1_000_000, 1)]

    async def fake_read(fn, *_args, **_kwargs):
        return rows if fn is tw.fetch_keyword_performance_rows else "UAH"

    monkeypatch.setattr(tw, "run_ads_read_call", fake_read)
    monkeypatch.setattr("mcp_server.artifacts.artifact_path", lambda _suffix: tmp_path / "out.xlsx")
    monkeypatch.setattr(
        "mcp_server.artifacts.publish_artifact",
        lambda _path, **kwargs: {"filename": kwargs["filename"], "token": "signed"},
    )

    result = asyncio.run(
        tw.export_keyword_report(
            DRAFT_ACCOUNT_ID, date_from="2025-01-01", date_to="2025-12-31"
        )
    )

    assert result["ok"] is True
    assert result["row_count"] == 1
    assert result["artifact"]["filename"].startswith("aimash_keywords_")
    assert load_workbook(tmp_path / "out.xlsx")["Ключевые слова"].max_row == 2


async def _async_value(value):
    return value
