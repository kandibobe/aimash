"""Лист «Находки» в выгрузках (/export, /sheets) + engine-only здоровье в MCC-книге.

Инварианты (каждый — про честность или про деньги, не про красоту):
• проза листа = проза карточки /audit (один источник, audit.render) — расходиться нечему;
• экспорт это БУМАГА: ни колонки suggested_operation, ни кнопки «применить» (золотое правило №3 —
  бюджет/ставка меняются только прямой командой через confirm-гейт);
• полный gather_audit (≈23 чтения) — ТОЛЬКО на явную команду человека по АККАУНТУ: покампанийный
  экспорт и MCC-книга его не зовут (квота), сбой аудита не роняет выгрузку;
• ячейки клиента обезврежены (safe_row: имя кампании `=HYPERLINK(...)` — текст, не формула);
• EN-книга без кириллицы (RU-утечка в артефакте).
"""

from __future__ import annotations

import pathlib
import sys
from datetime import date
from types import SimpleNamespace

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402
from audit.engine import build_audit  # noqa: E402
from reports.findings import FINDINGS_FORMATS, FINDINGS_TITLE, MONEY_COL, findings_rows  # noqa: E402
from reports.period import last_n_days  # noqa: E402
from reports.queries import Breakdown, Metrics  # noqa: E402
from reports.sheets import build_sheets_data, format_requests  # noqa: E402
from reports.xlsx import build_mcc_deep_workbook, build_workbook  # noqa: E402

ROOT = pathlib.Path(__file__).resolve().parents[1]
EVIL = '=HYPERLINK("http://evil","click")'
_CYR = set("абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ")


def _report(campaign: str = "Search Brand"):
    """Аккаунт с расходом и БЕЗ конверсий — движку есть что сказать (иначе тест зелен вхолостую)."""
    m = Metrics(impressions=1000, clicks=100, cost_micros=500_000_000, conversions=0.0)
    camp = Breakdown(
        key="campaign",
        title="Кампании",
        dim_headers=["Кампания", "Статус"],
        rows=[((campaign, "ENABLED"), m)],
    )
    return SimpleNamespace(
        customer_id="7753643025",
        totals=m,
        prev_totals=m,
        period=last_n_days(7, today=date(2026, 6, 25)),
        currency="USD",
        breakdowns=[camp],
    )


def _texts(wb, sheet: str) -> list[str]:
    ws = wb[sheet]
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


def test_findings_sheet_speaks_the_same_prose_as_the_audit_card():
    """Лист — не пересказ карточки, а она сама: обзор из render_audit + строка на КАЖДУЮ находку
    (не топ-8, как в чате: в этом и смысл выгрузки), в том же порядке worst-first."""
    from audit.render import finding_text, render_audit

    report = _report()
    result = build_audit(report)
    assert result.findings, "движку есть что сказать про этот аккаунт"

    wb = build_workbook(report, "ru", audit=result)
    assert wb.sheetnames[1] == FINDINGS_TITLE  # сразу за «Сводкой», перед разбивками
    texts = _texts(wb, FINDINGS_TITLE)
    for line in render_audit(result, "ru", actions=False).split("\n"):
        if line:
            assert line in texts, f"обзор карточки не доехал до листа: {line}"
    for f in result.findings:
        assert finding_text(f, "ru", result.currency) in texts
        assert f.check_id in texts
    rows = findings_rows(result, "ru")
    assert [r[-1] for r in rows] == [f.check_id for f in result.findings]  # порядок карточки


def test_findings_sheet_is_paper_not_a_button():
    """Золотое правило №3: экспорт не предлагает применить. Ни поля операции, ни кнопки — по коду."""
    src = (ROOT / "reports" / "findings.py").read_text(encoding="utf-8")
    assert "suggested_operation" not in src
    assert "ONE_TAP_OPS" not in src and "one_tap" not in src
    result = build_audit(_report())
    ops = {getattr(f, "suggested_operation", None) for f in result.findings}
    assert ops - {None}, "фикстура без one_tap-находок не проверяет запрет (нужна хоть одна)"
    for row in findings_rows(result, "ru"):
        assert not (set(map(str, row)) & (ops - {None}))


def test_findings_cells_are_formula_safe(tmp_path):
    """Имя кампании из Google (`=HYPERLINK…`) попадает в ячейку находки — обезврежено safe_row.
    Опасно только ВЕДУЩЕЕ =/+/-/@ (внутри прозы «Пауза «=HYPERLINK…»» формула не исполняется), поэтому
    гард двойной: ни одна ячейка листа не НАЧИНАЕТСЯ с лида, и колонка «Кампания» — с апострофом."""
    report = _report(campaign=EVIL)
    result = build_audit(report)
    path = str(tmp_path / "rep.xlsx")
    build_workbook(report, "ru", audit=result).save(path)
    ws = openpyxl.load_workbook(path)[FINDINGS_TITLE]
    cells = [c for row in ws.iter_rows() for c in row if isinstance(c.value, str)]
    for c in cells:
        assert not c.value.startswith(("=", "+", "-", "@")), f"формула в ячейке: {c.value!r}"
    named = [c for c in cells if c.value.lstrip("'") == EVIL]
    assert named, "имя кампании не доехало до листа — тест бесполезен"
    for c in named:
        assert c.data_type == "s" and c.value.startswith("'=")  # текст, НЕ формула


def test_workbook_without_audit_is_byte_for_byte_the_old_one():
    """audit=None (покампанийный экспорт, kill-switch, сбой сбора) → книга ровно как раньше."""
    wb = build_workbook(_report(), "ru")
    assert FINDINGS_TITLE not in wb.sheetnames
    assert wb.sheetnames == ["Сводка", "Кампании"]
    assert [t.title for t in build_sheets_data(_report(), "ru")] == ["Сводка", "Кампании"]


def test_findings_en_has_no_cyrillic():
    report = _report()
    result = build_audit(report)
    wb = build_workbook(report, "en", audit=result)
    leaks = [t for t in (_texts(wb, "Findings") + wb.sheetnames) if _CYR & set(t)]
    assert not leaks, f"RU-утечка в EN-книге (лист находок): {leaks}"
    tabs = build_sheets_data(report, "en", audit=result)
    cells = [str(c) for t in tabs for row in t.rows for c in row if c is not None]
    leaks = [t for t in cells + [t.title for t in tabs] if _CYR & set(t)]
    assert not leaks, f"RU-утечка во вкладке Sheets: {leaks}"


def test_sheets_findings_tab_formats_the_money_column_only():
    """У вкладки находок раскладка НЕ метрическая: форматы — строго по реестру FINDINGS_FORMATS
    («Под риском» + числовые факты), а не 9 колонок METRIC_FORMATS (иначе Sheets отформатировал бы
    прозу как проценты)."""
    report = _report()
    result = build_audit(report)
    tabs = build_sheets_data(report, "ru", audit=result)
    fin = next(t for t in tabs if t.title == FINDINGS_TITLE)
    assert fin.rows[fin.header_row][MONEY_COL].startswith("Под риском")
    num = [
        r["repeatCell"]
        for r in format_requests(tabs)
        if "repeatCell" in r
        and "numberFormat" in r["repeatCell"]["cell"]["userEnteredFormat"]
        and r["repeatCell"]["range"]["sheetId"] == tabs.index(fin)
    ]
    assert [rq["range"]["startColumnIndex"] for rq in num] == [c for c, _f in FINDINGS_FORMATS]
    for rq in num:
        assert rq["range"]["startRowIndex"] == fin.header_row + 1


def _child(cid, name):
    from ads.read import ChildAccount

    return ChildAccount(id=cid, name=name, currency="USD", manager=False, level=1, status="ENABLED")


def test_mcc_deep_health_is_engine_only():
    """Веер по MCC: балл/грейд/под-риском есть, а gather_audit (23 чтения × N аккаунтов) — нет.
    Гард структурный: слой reports вообще не вправе импортировать audit.collect."""
    for py in (ROOT / "reports").glob("*.py"):
        src = py.read_text(encoding="utf-8")
        assert "audit.collect" not in src and "gather_audit" not in src, (
            f"{py.name}: полный аудит в слое отчётов — это 23 чтения на КАЖДЫЙ аккаунт веера"
        )

    report = _report()
    deep = SimpleNamespace(
        manager_id="9990001112",
        period=report.period,
        items=[(_child("1111111111", "Tower"), report)],
        skipped=[],
        managers=[],
        inactive=[],
        errors=[],
    )
    wb = build_mcc_deep_workbook(deep, "ru")
    texts = _texts(wb, "Сводка")
    result = build_audit(report)
    assert "Балл" in texts and "Оценка" in texts
    assert str(result.score) in texts and result.grade in texts
    assert any("/audit" in t for t in texts)  # сноска: это НЕ полный аудит


def test_mcc_deep_survives_a_broken_report():
    """Движок споткнулся на одном аккаунте → пустые ячейки, а не упавшая книга."""
    from reports.findings import account_health_cells

    assert account_health_cells(SimpleNamespace()) == ["", "", ""]
