"""A9: экспорты (.xlsx/.csv/Sheets) не превращают данные клиента в формулы (CSV injection).

Имя кампании/ключа/аккаунта из Google Ads вроде `=HYPERLINK("http://evil","x")` попадает в ячейку.
openpyxl/Sheets трактуют ведущий `= + - @` как ФОРМУЛУ — reports.safe_cell префиксует апостроф,
и ячейка остаётся ТЕКСТОМ. Sheets-экспорт защищён режимом RAW (не парсит формулы). Гарды:
функционально (data_type ячейки == 's'/текст) + класс-греп (нет сырых append/writerow; Sheets=RAW).
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402

from reports.safe_cell import safe_cell, safe_row  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
EVIL = '=HYPERLINK("http://evil","click")'


# ── unit: safe_cell ───────────────────────────────────────────────────────────────
def test_safe_cell_prefixes_dangerous_leads():
    for lead in ("=", "+", "-", "@", "\t", "\r"):
        assert safe_cell(lead + "x") == "'" + lead + "x"


def test_safe_cell_leaves_safe_values_untouched():
    assert safe_cell("Бренд") == "Бренд"
    assert safe_cell("http://x.io") == "http://x.io"
    assert safe_cell(42) == 42  # число не трогаем (формат сохраняется)
    assert safe_cell(3.14) == 3.14
    assert safe_cell(None) is None


def test_safe_row_maps_each_cell():
    assert safe_row([EVIL, 10, "ok"]) == ["'" + EVIL, 10, "ok"]
    assert safe_row(None) == []


# ── функционально: ячейка с формулой становится текстом ───────────────────────────
def _first_cell_with(wb_path: str, needle: str):
    wb = openpyxl.load_workbook(wb_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and needle in c.value:
                    return c
    return None


def test_keyword_list_xlsx_neutralizes_formula(tmp_path):
    from keywords.export import write_keyword_list_xlsx

    path = str(tmp_path / "kw.xlsx")
    write_keyword_list_xlsx([EVIL, "нормальный ключ"], "phrase", "Добавить", path)
    cell = _first_cell_with(path, "HYPERLINK")
    assert cell is not None
    assert cell.data_type == "s"  # строка, НЕ формула ('f')
    assert cell.value.startswith("'=")  # обезврежено апострофом


def test_report_breakdown_xlsx_neutralizes_formula_in_dim(tmp_path):
    from openpyxl import Workbook

    from reports.queries import Breakdown, Metrics
    from reports.xlsx import _write_breakdown

    m = Metrics(impressions=1, clicks=1, cost_micros=1_000_000)
    b = Breakdown(key="campaign", title="Кампании", dim_headers=["Кампания"], rows=[((EVIL,), m)])
    wb = Workbook()
    _write_breakdown(wb, b, "USD")
    path = str(tmp_path / "rep.xlsx")
    wb.save(path)
    cell = _first_cell_with(path, "HYPERLINK")
    assert cell is not None and cell.data_type == "s"


def test_keywords_csv_neutralizes_formula(tmp_path):
    from keywords.cluster import Cluster
    from keywords.export import write_keywords_csv

    idea = _mk_idea(EVIL)
    cl = Cluster(name="Общий", intent="informational", keywords=[EVIL])
    path = str(tmp_path / "kw.csv")
    write_keywords_csv([cl], [idea], path)
    text = Path(path).read_text(encoding="utf-8-sig")
    assert "'=HYPERLINK" in text  # апостроф-префикс в CSV
    assert re.search(r"(^|,|\")=HYPERLINK", text) is None  # нет сырой формулы в начале поля


def _mk_idea(text: str):
    from ads.keyword_plan import KeywordIdea

    # минимальный KeywordIdea; поля объёма/ставок — дефолтные (не важны для теста инъекции)
    return KeywordIdea(text=text, avg_monthly_searches=10, competition="LOW")


# ── класс-гард: нет сырых append/writerow; Sheets=RAW ─────────────────────────────
def test_no_raw_append_or_writerow_in_exporters():
    for rel in ("reports/xlsx.py", "keywords/export.py"):
        src = (ROOT / rel).read_text(encoding="utf-8")
        # разрешён ТОЛЬКО append/writerow внутри helper-обёрток _append/_writerow
        for i, line in enumerate(src.splitlines(), 1):
            s = line.strip()
            if re.search(r"\b\w+\.append\(", s) and "def _append" not in s:
                # допускаем ТОЛЬКО ws.append(safe_row(...)) внутри _append
                assert "safe_row" in s, f"{rel}:{i}: сырой append мимо safe_row: {s}"
            if re.search(r"\b\w+\.writerow\(", s) and "def _writerow" not in s:
                assert "safe_row" in s, f"{rel}:{i}: сырой writerow мимо safe_row: {s}"


def test_sheets_uses_raw_input_option():
    """Sheets-экспорт защищён RAW (не парсит формулы). USER_ENTERED вернул бы уязвимость."""
    src = (ROOT / "reports/sheets.py").read_text(encoding="utf-8")
    assert "values().batchUpdate" in src
    assert "USER_ENTERED" not in src  # опасный режим не используется
    assert src.count('"valueInputOption": "RAW"') >= 2  # оба value-write под RAW
