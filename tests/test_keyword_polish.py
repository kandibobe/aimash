"""D8 (удобство 2026-07): полировка keyword research.

(1) индекс конкуренции (0–100) в чат-сводке — точнее, чем low/med/high;
(2) валюта аккаунта в денежных колонках .xlsx/.csv — ставки в micros→валюту считает КОД,
    но раньше колонки не подписывали валютой (не-USD аккаунт вводил в заблуждение).
"""

from __future__ import annotations

import csv
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import core.texts as texts  # noqa: E402
from ads.keyword_plan import KeywordIdea  # noqa: E402
from keywords.cluster import Cluster  # noqa: E402
from keywords.export import HEADERS, _headers, write_keywords_csv, write_keywords_xlsx  # noqa: E402


def test_competition_index_in_chat_suffix():
    idea = KeywordIdea(text="k", avg_monthly_searches=100, competition="HIGH", competition_index=85)
    suffix = texts._kw_metrics_suffix(idea, "USD", en=False)
    assert "(85)" in suffix  # индекс показан рядом с уровнем


def test_competition_index_hidden_when_zero():
    idea = KeywordIdea(text="k", competition="LOW", competition_index=0)
    suffix = texts._kw_metrics_suffix(idea, "", en=False)
    assert "(" not in suffix  # 0 (нет данных, тест-аккаунт) — не засоряем «(0)»


def test_headers_carry_currency_on_money_columns():
    h = _headers("JPY")
    money = [h[6], h[7], h[8]]  # Сред. CPC / Ставка низ / Ставка верх
    assert all("(JPY)" in x for x in money)
    # без валюты — как раньше (обратная совместимость)
    assert HEADERS[6] == "Сред. CPC" and "(" not in HEADERS[6]


def test_csv_export_header_has_currency():
    clusters = [Cluster(name="C", intent="brand", keywords=["k1"])]
    ideas = [KeywordIdea(text="k1", avg_monthly_searches=50, low_bid=1.0, high_bid=2.0)]
    with tempfile.TemporaryDirectory() as d:
        p = str(Path(d) / "k.csv")
        write_keywords_csv(clusters, ideas, p, seeds=["s"], currency="AUD")
        with open(p, encoding="utf-8-sig") as f:
            header = next(csv.reader(f))
    assert any("(AUD)" in col for col in header)


def test_xlsx_export_header_has_currency():
    from openpyxl import load_workbook

    clusters = [Cluster(name="C", intent="brand", keywords=["k1"])]
    ideas = [KeywordIdea(text="k1", avg_monthly_searches=50, low_bid=1.0, high_bid=2.0)]
    with tempfile.TemporaryDirectory() as d:
        p = str(Path(d) / "k.xlsx")
        write_keywords_xlsx(clusters, ideas, p, seeds=["s"], currency="UAH")
        wb = load_workbook(p)
        ws = wb.active
        # шапка на строке 4 (2 строки мета + пустая); ищем колонку ставки с валютой
        row4 = [c.value for c in ws[4]]
    assert any(v and "(UAH)" in str(v) for v in row4)
