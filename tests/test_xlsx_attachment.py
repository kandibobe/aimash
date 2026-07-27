"""Тесты confirm/xlsx_attachment.py: генерация .xlsx-вложения для больших списков ключей."""

from __future__ import annotations

import base64
import os
import sys
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import load_workbook

from confirm.xlsx_attachment import write_keywords_xlsx


# ── Мок тяжёлых импортов (bare-host без google-ads/asyncpg/aiogram/…) ──

class _FakeModule(MagicMock):
    """Рекурсивный мок для отсутствующих C++/Docker-only пакетов."""

_FAKE = _FakeModule()

# google-ads SDK (C++ protobuf, ~200 MB)
for _mod in (
    "google", "google.ads", "google.ads.googleads",
    "google.ads.googleads.client", "google.ads.googleads.errors",
    "google.ads.googleads.v17", "google.ads.googleads.v19",
    "google.api_core", "google.protobuf", "google.type",
    "google.rpc", "google.longrunning",
):
    sys.modules.setdefault(_mod, _FAKE)

# Прочие тяжёлые / отсутствующие на bare-host
for _mod in (
    "aiogram", "asyncpg", "tenacity", "cachetools", "langfuse",
    "grpc", "proto", "protobuf", "googleads", "openai",
    "tiktoken", "sentry_sdk", "apscheduler", "aiohttp",
    "psycopg2", "psycopg", "redis", "celery", "kombu",
    "billiard", "vine", "amqp",
):
    sys.modules.setdefault(_mod, _FAKE)

# db/session.py вызывает create_async_engine при импорте — не даём ему
# попытаться подключиться к БД (asyncpg не установлен, да и не нужен тестам).
_sa_mock = MagicMock()
_sa_mock.ext.asyncio.create_async_engine = MagicMock(return_value=MagicMock())
_sa_mock.ext.asyncio.async_sessionmaker = MagicMock(return_value=MagicMock())
sys.modules.setdefault("sqlalchemy.ext.asyncio", _sa_mock.ext.asyncio)


# ── Тесты write_keywords_xlsx ──────────────────────────────────────────


def test_generates_xlsx_with_30_keywords():
    """30 ключей → .xlsx существует и открывается openpyxl."""
    keywords = [f"ключевое слово {i}" for i in range(1, 31)]
    path = write_keywords_xlsx(
        keywords=keywords,
        operation="add_keywords",
        match_type="широкое",
        campaign="Test Campaign",
    )
    assert os.path.isfile(path), f"файл не создан: {path}"
    try:
        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 31, f"ожидалось 31 строка, получено {ws.max_row}"
        assert ws.max_column == 4, f"ожидалось 4 колонки, получено {ws.max_column}"
        wb.close()
    finally:
        tmpdir = os.path.dirname(path)
        if os.path.isfile(path):
            os.remove(path)
        if os.path.isdir(tmpdir):
            try:
                os.rmdir(tmpdir)
            except OSError:
                pass


def test_generates_xlsx_with_few_keywords():
    """5 ключей — функция ВСЁ РАВНО генерирует файл (порог — в build_proposal)."""
    keywords = [f"kw{i}" for i in range(5)]
    path = write_keywords_xlsx(
        keywords=keywords,
        operation="add_negative_keywords",
        match_type="точное",
    )
    assert os.path.isfile(path)
    try:
        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 6
        wb.close()
    finally:
        tmpdir = os.path.dirname(path)
        if os.path.isfile(path):
            os.remove(path)
        if os.path.isdir(tmpdir):
            try:
                os.rmdir(tmpdir)
            except OSError:
                pass


def test_header_columns():
    """Заголовки колонок — «Ключевое слово», «Тип соответствия», «Кампания», «Операция»."""
    keywords = ["test keyword"]
    path = write_keywords_xlsx(
        keywords=keywords,
        operation="remove_keywords",
        match_type="фразовое",
        campaign="MyCampaign",
    )
    try:
        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        expected = ["Ключевое слово", "Тип соответствия", "Кампания", "Операция"]
        assert headers == expected, f"заголовки: {headers}, ожидались: {expected}"
        wb.close()
    finally:
        tmpdir = os.path.dirname(path)
        if os.path.isfile(path):
            os.remove(path)
        if os.path.isdir(tmpdir):
            try:
                os.rmdir(tmpdir)
            except OSError:
                pass


def test_header_style():
    """Шапка — жирный белый шрифт на синем фоне #4472C4."""
    keywords = ["kw"]
    path = write_keywords_xlsx(keywords=keywords, operation="add_keywords")
    try:
        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None
        for c in range(1, 5):
            cell = ws.cell(row=1, column=c)
            assert cell.font.bold is True, f"колонка {c}: ожидался жирный шрифт"
            font_color = str(cell.font.color.rgb) if cell.font.color else ""
            assert "FFFFFF" in font_color.upper(), (
                f"колонка {c}: ожидался белый шрифт, получен {font_color!r}"
            )
            fill_fg = str(cell.fill.fgColor.rgb) if cell.fill.fgColor else ""
            assert "4472C4" in fill_fg.upper(), (
                f"колонка {c}: ожидался синий фон #4472C4, получен {fill_fg!r}"
            )
        wb.close()
    finally:
        tmpdir = os.path.dirname(path)
        if os.path.isfile(path):
            os.remove(path)
        if os.path.isdir(tmpdir):
            try:
                os.rmdir(tmpdir)
            except OSError:
                pass


def test_data_rows_contain_correct_values():
    """Строки данных содержат ключ, тип, кампанию и операцию."""
    keywords = ["красные розы", "белые тюльпаны"]
    path = write_keywords_xlsx(
        keywords=keywords, operation="add_keywords",
        match_type="широкое", campaign="Flowers",
    )
    try:
        wb = load_workbook(path)
        ws = wb.active
        assert ws is not None
        row2 = [ws.cell(row=2, column=c).value for c in range(1, 5)]
        assert row2 == ["красные розы", "широкое", "Flowers", "add_keywords"]
        row3 = [ws.cell(row=3, column=c).value for c in range(1, 5)]
        assert row3 == ["белые тюльпаны", "широкое", "Flowers", "add_keywords"]
        wb.close()
    finally:
        tmpdir = os.path.dirname(path)
        if os.path.isfile(path):
            os.remove(path)
        if os.path.isdir(tmpdir):
            try:
                os.rmdir(tmpdir)
            except OSError:
                pass


# ── Интеграция: build_proposal ─────────────────────────────────────────


@pytest.mark.asyncio
async def test_build_proposal_skips_xlsx_for_few_keywords():
    """5 ключей → build_proposal НЕ генерирует xlsx."""
    from bot.proposal import build_proposal

    store = MagicMock()
    store.save_proposal = AsyncMock()

    with (
        patch("bot.proposal.read_state", new_callable=AsyncMock) as mock_read,
        patch("bot.proposal.attach_freshness", side_effect=lambda p, s: p),
    ):
        mock_read.return_value = MagicMock()
        result = await build_proposal(
            store=store,
            operation="add_keywords",
            params={"keywords": [f"kw{i}" for i in range(5)], "match_type": "broad", "campaign": "Test"},
            cid="test-cid-123",
            chat_id=12345,
        )

    assert result.big_list_attachment is None


@pytest.mark.asyncio
async def test_build_proposal_generates_xlsx_for_many_keywords():
    """30 ключей → build_proposal генерирует xlsx."""
    from bot.proposal import build_proposal

    store = MagicMock()
    store.save_proposal = AsyncMock()

    with (
        patch("bot.proposal.read_state", new_callable=AsyncMock) as mock_read,
        patch("bot.proposal.attach_freshness", side_effect=lambda p, s: p),
    ):
        mock_read.return_value = MagicMock()
        result = await build_proposal(
            store=store,
            operation="add_keywords",
            params={"keywords": [f"kw{i}" for i in range(30)], "match_type": "broad", "campaign": "Test"},
            cid="test-cid-456",
            chat_id=12345,
        )

    try:
        assert result.big_list_attachment is not None
        assert os.path.isfile(result.big_list_attachment)
        wb = load_workbook(result.big_list_attachment)
        ws = wb.active
        assert ws is not None
        assert ws.max_row == 31
        wb.close()
    finally:
        if result.big_list_attachment and os.path.isfile(result.big_list_attachment):
            tmpdir = os.path.dirname(result.big_list_attachment)
            os.remove(result.big_list_attachment)
            try:
                os.rmdir(tmpdir)
            except OSError:
                pass


@pytest.mark.asyncio
async def test_build_proposal_xlsx_for_negative_shared_set():
    """add_negatives_to_shared_set с >20 ключей — тоже генерирует xlsx."""
    from bot.proposal import build_proposal

    store = MagicMock()
    store.save_proposal = AsyncMock()

    with (
        patch("bot.proposal.read_state", new_callable=AsyncMock) as mock_read,
        patch("bot.proposal.attach_freshness", side_effect=lambda p, s: p),
    ):
        mock_read.return_value = MagicMock()
        result = await build_proposal(
            store=store,
            operation="add_negatives_to_shared_set",
            params={"keywords": [f"neg{i}" for i in range(25)], "match_type": "exact", "shared_set": "Global Negatives"},
            cid="test-cid-789",
            chat_id=12345,
        )

    try:
        assert result.big_list_attachment is not None
        assert os.path.isfile(result.big_list_attachment)
        wb = load_workbook(result.big_list_attachment)
        ws = wb.active
        assert ws is not None
        camp_col = ws.cell(row=2, column=3).value
        assert camp_col == "Global Negatives", f"ожидалась 'Global Negatives', получено {camp_col!r}"
        wb.close()
    finally:
        if result.big_list_attachment and os.path.isfile(result.big_list_attachment):
            tmpdir = os.path.dirname(result.big_list_attachment)
            os.remove(result.big_list_attachment)
            try:
                os.rmdir(tmpdir)
            except OSError:
                pass


# ── cleanup_attachment ─────────────────────────────────────────────────


def test_cleanup_attachment_removes_file_and_dir() -> None:
    """cleanup_attachment удаляет .xlsx-файл и его временную директорию."""
    from confirm.xlsx_attachment import cleanup_attachment

    path = write_keywords_xlsx(keywords=["test"], operation="add_keywords")
    tmpdir = os.path.dirname(path)
    assert os.path.isfile(path)
    assert os.path.isdir(tmpdir)

    cleanup_attachment(path)

    assert not os.path.isfile(path)
    assert not os.path.isdir(tmpdir)


def test_cleanup_attachment_idempotent() -> None:
    """Повторный вызов cleanup_attachment не падает."""
    from confirm.xlsx_attachment import cleanup_attachment

    path = write_keywords_xlsx(keywords=["test"], operation="add_keywords")
    cleanup_attachment(path)
    cleanup_attachment(path)


def test_cleanup_attachment_on_empty_path() -> None:
    """cleanup_attachment('') / несуществующий путь не падает."""
    from confirm.xlsx_attachment import cleanup_attachment

    cleanup_attachment("")
    cleanup_attachment("/nonexistent/path/file.xlsx")


# ── read_attachment_b64 ────────────────────────────────────────────────


def test_read_attachment_b64_returns_valid_base64() -> None:
    """read_attachment_b64 возвращает base64-строку, декодируемую обратно в .xlsx."""
    from confirm.xlsx_attachment import cleanup_attachment, read_attachment_b64

    path = write_keywords_xlsx(keywords=["kw1", "kw2"], operation="remove_keywords")
    try:
        b64 = read_attachment_b64(path)
        assert isinstance(b64, str)
        raw = base64.b64decode(b64)
        assert raw[:2] == b"PK", f"ожидалась ZIP-сигнатура PK, получено {raw[:4]!r}"
        assert len(raw) > 0
    finally:
        cleanup_attachment(path)


# ── Интеграция: MCP _propose ───────────────────────────────────────────


@pytest.mark.asyncio
async def test_propose_includes_attachment_for_many_keywords() -> None:
    """MCP _propose с 30 ключами → attachment в строке конверта."""
    from mcp_server.tools_writes import _propose

    store = MagicMock()
    store.save_proposal = AsyncMock()

    with (
        patch("mcp_server.tools_writes._get_store", return_value=store),
        patch("bot.proposal.read_state", new_callable=AsyncMock) as mock_read,
        patch("bot.proposal.attach_freshness", side_effect=lambda p, s: p),
    ):
        mock_read.return_value = MagicMock()
        result = await _propose(
            operation="add_keywords",
            params={"keywords": [f"kw{i}" for i in range(30)], "match_type": "broad", "campaign": "Test"},
            account="7753643025",
        )

    assert "rows" in result
    assert len(result["rows"]) == 1
    row = result["rows"][0]
    assert "attachment" in row, f"для 30 ключей ожидался attachment, ключи строки: {list(row.keys())}"
    att = row["attachment"]
    assert "content_b64" in att
    assert "filename" in att
    assert att["filename"].endswith(".xlsx")
    raw = base64.b64decode(att["content_b64"])
    assert raw[:2] == b"PK"


@pytest.mark.asyncio
async def test_propose_no_attachment_for_few_keywords() -> None:
    """MCP _propose с 5 ключами → attachment НЕ включается."""
    from mcp_server.tools_writes import _propose

    store = MagicMock()
    store.save_proposal = AsyncMock()

    with (
        patch("mcp_server.tools_writes._get_store", return_value=store),
        patch("bot.proposal.read_state", new_callable=AsyncMock) as mock_read,
        patch("bot.proposal.attach_freshness", side_effect=lambda p, s: p),
    ):
        mock_read.return_value = MagicMock()
        result = await _propose(
            operation="add_keywords",
            params={"keywords": [f"kw{i}" for i in range(5)], "match_type": "broad", "campaign": "Test"},
            account="7753643025",
        )

    row = result["rows"][0]
    assert "attachment" not in row, f"для 5 ключей attachment не должен присутствовать, но получен"