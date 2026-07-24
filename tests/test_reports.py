"""Офлайн-тесты Фазы 2.A: глубокий отчёт по аккаунту + .xlsx. READ-ONLY, без живого Google Ads.

SDK подменяется фейковым клиентом (get_service→search возвращает заранее заданные строки).
Проверяем: математику периода, производные метрики (CTR/CPC/CPA/ROAS), агрегацию из строк,
замок аккаунта на чтении (golden rule #9), структуру .xlsx.
"""

from __future__ import annotations

import sys
from contextlib import contextmanager
from datetime import date
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from ads.client import DRAFT_ACCOUNT_ID  # noqa: E402
from core.config import settings  # noqa: E402
from reports import period as P  # noqa: E402
from reports import queries as Q  # noqa: E402
from reports import service as S  # noqa: E402
from reports import xlsx as X  # noqa: E402


@contextmanager
def allowed_ids(value: str):
    prev = settings.google_ads_allowed_customer_ids
    settings.google_ads_allowed_customer_ids = value
    try:
        yield
    finally:
        settings.google_ads_allowed_customer_ids = prev


# ── Фейковый Google Ads клиент: одна «универсальная» строка под любой запрос ──────
def _row(cost_micros: int, name: str = "Camp"):
    return SimpleNamespace(
        metrics=SimpleNamespace(
            impressions=100,
            clicks=10,
            cost_micros=cost_micros,
            conversions=2.0,
            conversions_value=50.0,
        ),
        campaign=SimpleNamespace(name=name, status=SimpleNamespace(name="ENABLED")),
        ad_group=SimpleNamespace(name="AG", status=SimpleNamespace(name="ENABLED")),
        ad_group_criterion=SimpleNamespace(
            keyword=SimpleNamespace(text="купить цветы", match_type=SimpleNamespace(name="PHRASE"))
        ),
        ad_group_ad=SimpleNamespace(
            ad=SimpleNamespace(id=123, type=SimpleNamespace(name="RESPONSIVE_SEARCH_AD"))
        ),
        segments=SimpleNamespace(
            device=SimpleNamespace(name="MOBILE"),
            ad_network_type=SimpleNamespace(name="SEARCH"),
            date="2026-06-01",
        ),
    )


class _FakeGA:
    def __init__(self, rows):
        self._rows = rows

    def search(self, *, customer_id, query):
        return list(self._rows)


class _FakeClient:
    def __init__(self, rows):
        self._rows = rows

    def get_service(self, name):
        assert name == "GoogleAdsService"
        return _FakeGA(self._rows)


# ── Период ───────────────────────────────────────────────────────────────────────
def test_last_n_days_excludes_today():
    today = date(2026, 6, 25)
    p = P.last_n_days(30, today=today)
    assert p.date_to == date(2026, 6, 24)  # вчера
    assert p.date_from == date(2026, 5, 26)
    assert p.days == 30
    assert p.gaql_between() == "segments.date BETWEEN '2026-05-26' AND '2026-06-24'"


def test_previous_period_is_equal_length_and_adjacent():
    p = P.last_n_days(7, today=date(2026, 6, 25))
    prev = p.previous()
    assert prev.date_to == date(2026, 6, 17)  # день перед началом текущего
    assert prev.days == 7


def test_month_to_date_and_first_of_month_edge():
    assert P.month_to_date(today=date(2026, 6, 25)).date_from == date(2026, 6, 1)
    p1 = P.month_to_date(today=date(2026, 6, 1))  # 1-е число: полных дней нет
    assert p1.date_from == date(2026, 6, 1) and p1.date_to == date(2026, 6, 1)


def test_from_preset_and_custom():
    assert P.from_preset("90", today=date(2026, 6, 25)).days == 90
    assert P.from_preset("14", today=date(2026, 6, 25)).days == 14  # 3.1: новый пресет
    assert P.from_preset("MTD", today=date(2026, 6, 25)).date_from == date(2026, 6, 1)
    assert P.custom(date(2026, 1, 1), date(2026, 1, 31)).days == 31
    for bad in ("5", "year", ""):
        try:
            P.from_preset(bad)
            raise AssertionError(f"должно было упасть: {bad!r}")
        except ValueError:
            pass
    try:
        P.custom(date(2026, 2, 2), date(2026, 2, 1))
        raise AssertionError("date_to < date_from должно падать")
    except ValueError:
        pass


def test_last_month_preset_and_reanchor():
    """3.1: пресет LM = прошлый КАЛЕНДАРНЫЙ месяц; kind='last_month' — окно относительное, поэтому
    reports.tz.reanchor пере-якорит его на «сегодня» аккаунта (граница месяца зависит от TZ)."""
    lm = P.from_preset("LM", today=date(2026, 6, 25))
    assert lm.kind == "last_month"
    assert lm.date_from == date(2026, 5, 1) and lm.date_to == date(2026, 5, 31)
    # январь → прошлый месяц = декабрь ПРОШЛОГО года (граница года)
    jan = P.last_month(today=date(2026, 1, 5))
    assert jan.date_from == date(2025, 12, 1) and jan.date_to == date(2025, 12, 31)
    # reanchor: «сегодня» аккаунта в другом месяце → окно съезжает на ЕГО прошлый месяц
    from reports.tz import reanchor

    re_lm = reanchor(lm, today=date(2026, 7, 2))
    assert re_lm.date_from == date(2026, 6, 1) and re_lm.date_to == date(2026, 6, 30)
    # custom остаётся как есть (абсолютные даты не пере-якорим)
    cust = P.custom(date(2026, 1, 1), date(2026, 1, 31))
    assert reanchor(cust, today=date(2026, 7, 2)) is cust


# ── Метрики (производные считает КОД) ────────────────────────────────────────────
def test_metrics_derived_and_sum():
    m = Q.Metrics(
        impressions=1000, clicks=50, cost_micros=25_000_000, conversions=5.0, conv_value=200.0
    )
    assert m.cost == 25.0
    assert abs(m.ctr - 0.05) < 1e-9
    assert m.avg_cpc == 0.5
    assert m.cpa == 5.0
    assert m.roas == 8.0
    # деление на ноль безопасно
    z = Q.Metrics()
    assert z.ctr == 0.0 and z.avg_cpc == 0.0 and z.cpa == 0.0 and z.roas == 0.0
    z.add(m)
    assert z.clicks == 50 and z.cost_micros == 25_000_000
    assert len(m.as_row()) == len(Q.METRIC_HEADERS)


# ── Замок аккаунта на чтении (golden rule #9) ────────────────────────────────────
def test_fetchers_reject_foreign_account():
    p = P.last_n_days(7, today=date(2026, 6, 25))
    client = _FakeClient([_row(1_000_000)])
    with allowed_ids(DRAFT_ACCOUNT_ID):
        for fn in (Q.fetch_totals, *Q.BREAKDOWN_FETCHERS):
            try:
                fn(client, "1234567890", p)
                raise AssertionError(f"{fn.__name__}: чужой аккаунт должен падать")
            except PermissionError:
                pass


# ── Агрегация из строк ───────────────────────────────────────────────────────────
def test_fetch_totals_and_breakdowns_aggregate():
    p = P.last_n_days(7, today=date(2026, 6, 25))
    client = _FakeClient([_row(3_000_000, "A"), _row(1_000_000, "B")])
    with allowed_ids(DRAFT_ACCOUNT_ID):
        totals = Q.fetch_totals(client, DRAFT_ACCOUNT_ID, p)
        assert totals.cost == 4.0 and totals.clicks == 20  # сумма двух строк
        camp = Q.fetch_by_campaign(client, DRAFT_ACCOUNT_ID, p)
        assert camp.key == "campaign" and len(camp.rows) == 2
        assert camp.dim_headers == ["Кампания", "Статус"]
        kw = Q.fetch_by_keyword(client, DRAFT_ACCOUNT_ID, p)
        assert kw.rows[0][0][2] == "купить цветы"  # текст ключа
        ad = Q.fetch_by_ad(client, DRAFT_ACCOUNT_ID, p)
        assert ad.rows[0][0][3] == "RESPONSIVE_SEARCH_AD"  # ad.type.name
        dev = Q.fetch_by_device(client, DRAFT_ACCOUNT_ID, p)
        assert dev.rows[0][0][0] == "MOBILE"


# ── Сборка отчёта + текстовая сводка ─────────────────────────────────────────────
def test_build_account_report_and_summary():
    p = P.last_n_days(7, today=date(2026, 6, 25))
    client = _FakeClient([_row(2_000_000, "A")])
    with allowed_ids(DRAFT_ACCOUNT_ID):
        report = S.build_account_report(client, DRAFT_ACCOUNT_ID, p, with_comparison=True)
    assert report.prev_totals is not None
    assert len(report.breakdowns) == len(Q.BREAKDOWN_FETCHERS)
    text = S.summary_text(report)
    assert DRAFT_ACCOUNT_ID in text and "Расход" in text and "к пред. периоду" in text


def test_build_report_without_comparison():
    p = P.last_n_days(30, today=date(2026, 6, 25))
    client = _FakeClient([_row(1_000_000)])
    with allowed_ids(DRAFT_ACCOUNT_ID):
        report = S.build_account_report(client, DRAFT_ACCOUNT_ID, p, with_comparison=False)
    assert report.prev_totals is None


# ── Ф0: анти-ложноположительные фетчеры аудита (claude-ads accuracy notes) ───────
class _RoutedGA:
    """search() отдаёт разные строки по ресурсу в FROM — фетчеры аудита делают 2-3 запроса.
    Все запросы пишутся в `seen` (клиент), чтобы тест мог пинуть сам GAQL."""

    def __init__(self, by_resource: dict, seen: list):
        self._by = by_resource
        self._seen = seen

    def search(self, *, customer_id, query):
        self._seen.append(query)
        for resource, rows in self._by.items():
            if (
                f"FROM {resource}" in query
            ):  # «FROM shared_set» НЕ матчит «FROM campaign_shared_set»
                return list(rows)
        return []


class _RoutedClient:
    def __init__(self, by_resource: dict):
        self._by = by_resource
        self.seen: list[str] = []

    def get_service(self, name):
        assert name == "GoogleAdsService"
        return _RoutedGA(self._by, self.seen)


def test_normalize_keyword_text_dedups_match_types_and_bmm():
    """G03: один ключ в разных типах соответствия (и legacy BMM с «+») — ОДИН ключ, не три."""
    variants = ("купить обувь", "+купить +обувь", "Купить  Обувь")
    assert len({Q.normalize_keyword_text(t) for t in variants}) == 1
    assert Q.normalize_keyword_text("кроссовки") != Q.normalize_keyword_text("купить обувь")


def test_adgroup_structure_counts_unique_keyword_texts():
    """G03 accuracy note: счётчик «свалки» = УНИКАЛЬНЫЕ тексты ключей, а не строки keyword_view.
    Ключи БЕЗ показов отсекает сам GAQL (metrics.impressions > 0) — здесь пинуем дедуп."""
    p = P.last_n_days(7, today=date(2026, 6, 25))

    def kw(text):
        return SimpleNamespace(
            ad_group=SimpleNamespace(id=1, name="AG"),
            campaign=SimpleNamespace(name="Main"),
            ad_group_criterion=SimpleNamespace(keyword=SimpleNamespace(text=text)),
        )

    client = _RoutedClient(
        {
            "keyword_view": [
                kw("купить обувь"),
                kw("+купить +обувь"),
                kw("Купить Обувь"),
                kw("кроссовки"),
            ],
            "ad_group_ad": [
                SimpleNamespace(
                    ad_group=SimpleNamespace(id=1, name="AG"),
                    campaign=SimpleNamespace(name="Main"),
                    ad_group_ad=SimpleNamespace(
                        ad=SimpleNamespace(type_=SimpleNamespace(name="RESPONSIVE_SEARCH_AD"))
                    ),
                )
            ],
        }
    )
    with allowed_ids(DRAFT_ACCOUNT_ID):
        rows = Q.fetch_adgroup_structure(client, DRAFT_ACCOUNT_ID, p)
    assert len(rows) == 1
    assert rows[0].kw_count == 2  # не 4: три варианта одного ключа схлопнулись
    assert rows[0].rsa_count == 1
    # Спящий хвост отсекает сам GAQL: ключи берём из keyword_view (у ad_group_criterion метрик нет)
    kw_q = next(q for q in client.seen if "FROM keyword_view" in q)
    assert "metrics.impressions > 0" in kw_q and "ad_group.status = 'ENABLED'" in kw_q


def test_negative_lists_sees_campaign_level_negatives():
    """G14/G15: минусы ПРЯМО на кампании считаются наравне со shared-списком. Аккаунт без списков,
    но с минусами на кампании → `campaign_level_count > 0` и кампания в карте покрытия."""
    neg = SimpleNamespace(
        campaign=SimpleNamespace(name="Main"),
        campaign_criterion=SimpleNamespace(negative=True),
    )
    positive = SimpleNamespace(  # НЕ минус (обычный критерий) → в счёт не идёт
        campaign=SimpleNamespace(name="Other"),
        campaign_criterion=SimpleNamespace(negative=False),
    )
    client = _RoutedClient(
        {
            "shared_set": [],  # shared-списков нет вовсе
            "campaign_shared_set": [],
            "campaign_criterion": [neg, neg, positive],
        }
    )
    with allowed_ids(DRAFT_ACCOUNT_ID):
        info = Q.fetch_negative_lists(client, DRAFT_ACCOUNT_ID)
    assert info.count == 0 and info.campaign_level_count == 2
    assert info.campaigns_with_negatives == frozenset({"Main"})  # «Other» — не минус


def test_negative_lists_ignores_non_keyword_shared_sets():
    """Привязанный к кампании PLACEMENT_EXCLUSION-сет — НЕ минус-слова: в покрытие не идёт."""
    client = _RoutedClient(
        {
            "shared_set": [SimpleNamespace(shared_set=SimpleNamespace(id=1))],
            "campaign_shared_set": [
                SimpleNamespace(
                    campaign=SimpleNamespace(name="Main"),
                    shared_set=SimpleNamespace(type_=SimpleNamespace(name="NEGATIVE_KEYWORDS")),
                ),
                SimpleNamespace(
                    campaign=SimpleNamespace(name="Display"),
                    shared_set=SimpleNamespace(type_=SimpleNamespace(name="PLACEMENT_EXCLUSION")),
                ),
            ],
            "campaign_criterion": [],
        }
    )
    with allowed_ids(DRAFT_ACCOUNT_ID):
        info = Q.fetch_negative_lists(client, DRAFT_ACCOUNT_ID)
    assert info.count == 1 and info.attached_campaigns == 1
    assert info.campaigns_with_negatives == frozenset({"Main"})


# ── .xlsx: структура и формат ────────────────────────────────────────────────────
def test_xlsx_workbook_structure():
    from openpyxl import load_workbook

    p = P.last_n_days(7, today=date(2026, 6, 25))
    client = _FakeClient([_row(2_000_000, "A"), _row(1_000_000, "B")])
    with allowed_ids(DRAFT_ACCOUNT_ID):
        report = S.build_account_report(client, DRAFT_ACCOUNT_ID, p)
    wb = X.build_workbook(report)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)

    assert wb2.sheetnames[0] == "Сводка"
    for title in ("Кампании", "Группы объявлений", "Ключевые слова", "По дням"):
        assert title in wb2.sheetnames
    ws = wb2["Кампании"]
    headers = [c.value for c in ws[1]]
    assert headers == ["Кампания", "Статус", *Q.METRIC_HEADERS]
    # данные на месте: первая кампания A с расходом 2.0
    assert ws.cell(row=2, column=1).value == "A"
    assert ws.cell(row=2, column=7).value == 2.0  # колонка «Расход» (2 dim + 5-я метрика)
