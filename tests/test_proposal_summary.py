"""Тесты человекочитаемой сводки черновика мутации (ТЗ §5) + компактного .xlsx списка ключей.

Закрывает находку аудита: сводка add_keywords/минус-слов была сырым dict, без типа соответствия
словами и без XLSX для больших списков.
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook  # noqa: E402

from core import texts  # noqa: E402
from keywords.export import write_keyword_list_xlsx  # noqa: E402


def test_fmt_add_keywords_human_readable():
    s = texts.fmt_mutation_summary(
        "add_keywords",
        {"campaign": "Search", "keywords": ["купить телефон", "смартфон"], "match_type": "phrase"},
    )
    assert "Search" in s
    assert "фразовое" in s  # тип соответствия словами, не 'phrase'
    assert "купить телефон" in s and "смартфон" in s
    assert "{" not in s and "'" not in s  # не сырой dict/repr


def test_fmt_add_keywords_truncates_big_list_with_xlsx_note():
    kws = [f"kw{i}" for i in range(30)]
    s = texts.fmt_mutation_summary(
        "add_keywords", {"campaign": "X", "keywords": kws, "match_type": "broad"}
    )
    assert "широкое" in s
    assert s.count("•") == texts.KW_INLINE_MAX  # ровно KW_INLINE_MAX в тексте
    assert f"ещё {30 - texts.KW_INLINE_MAX}" in s and ".xlsx" in s


def test_fmt_negative_and_remove_labels():
    neg = texts.fmt_mutation_summary(
        "add_negative_keywords", {"campaign": "X", "keywords": ["spam"], "match_type": "broad"}
    )
    assert "минус-слов" in neg
    rm = texts.fmt_mutation_summary(
        "remove_keywords", {"campaign": "X", "keywords": ["old"], "match_type": "exact"}
    )
    assert "удалить" in rm.lower() and "точное" in rm


def test_fmt_budget_and_bid():
    assert "+20%" in texts.fmt_mutation_summary(
        "update_budget",
        {"campaign": "X", "mode": "increase_by_percent", "value": 20, "currency": "percent"},
    )
    s = texts.fmt_mutation_summary(
        "update_budget", {"campaign": "X", "mode": "set_to", "value": 50, "currency": "USD"}
    )
    assert "→ 50 USD" in s
    assert "ставка CPC" in texts.fmt_mutation_summary(
        "update_bid", {"campaign": "X", "mode": "set_to", "value": 1.5, "currency": "USD"}
    )


def test_fmt_pause_resume_geo():
    assert "паузу" in texts.fmt_mutation_summary("pause_campaign", {"campaign": "X"})
    assert "возобновить" in texts.fmt_mutation_summary("resume_campaign", {"campaign": "X"})
    g = texts.fmt_mutation_summary(
        "set_geo_proximity",
        {"campaign": "X", "radius_km": 5, "city_name": "Київ", "country_code": "UA"},
    )
    assert "5 км" in g and "Київ" in g


def test_fmt_returns_empty_for_rich_ops():
    # create_rsa/create_gdn имеют свой богатый форматтер → caller оставляет собственный summary
    assert texts.fmt_mutation_summary("create_rsa", {"campaign": "X"}) == ""
    assert texts.fmt_mutation_summary("create_gdn_campaign", {"campaign": "X"}) == ""


# ── §10: редакторские замечания видны на карточке ДО «да» (bot-free дом эвристики) ──
_CLEAN_H = ["Заголовок один", "Заголовок два", "Заголовок три"]
_CLEAN_D = ["Описание первое.", "Описание второе."]


def test_rsa_card_carries_moderation_advisory_ru():
    dirty = texts.fmt_rsa_proposal_summary(
        "Группа", ["OZON доставка", *_CLEAN_H[1:]], _CLEAN_D, "https://x.example/", lang="ru"
    )
    assert "⚠️" in dirty and "модерац" in dirty
    clean = texts.fmt_rsa_proposal_summary(
        "Группа", _CLEAN_H, _CLEAN_D, "https://x.example/", lang="ru"
    )
    assert "⚠️" not in clean  # чисто → карточка без шума


def test_rsa_card_carries_moderation_advisory_en():
    s = texts.fmt_rsa_proposal_summary(
        "Group",
        ["OZON delivery", "Headline two", "Headline three"],
        ["First description.", "Second description."],
        "https://x.example/",
        lang="en",
    )
    assert "⚠️" in s and "ad review" in s


def _search_card(headlines, lang="ru"):
    return texts.fmt_search_proposal_summary(
        "Кампания",
        "https://x.example/",
        100.0,
        headlines,
        _CLEAN_D,
        ["купить телефон"],
        "phrase",
        lang=lang,
    )


def test_search_campaign_card_carries_moderation_advisory():
    """Второй вызывающий _validate_rsa_inputs строит карточку ДРУГИМ форматтером — advisory
    обязан быть и там, иначе на пути create_search_campaign замечание доходит только в лог."""
    dirty = _search_card(["OZON доставка", *_CLEAN_H[1:]])
    assert "⚠️" in dirty and "модерац" in dirty
    assert "⚠️" not in _search_card(_CLEAN_H)
    assert "ad review" in _search_card(["OZON delivery", "Headline two"], lang="en")


def test_clone_card_inherits_moderation_advisory():
    """Сводка клона переиспользует тело fmt_search_proposal_summary — строка наследуется."""
    s = texts.fmt_clone_proposal_summary(
        "Новая",
        "Старая",
        100.0,
        {
            "final_url": "https://x.example/",
            "headlines": ["OZON доставка", *_CLEAN_H[1:]],
            "descriptions": _CLEAN_D,
            "keywords": ["купить телефон"],
            "match_type": "phrase",
        },
        lang="ru",
    )
    assert "⚠️" in s and "модерац" in s
    assert "Не переносится автоматически" in s  # хвост клона на месте, не срезан advisory


def test_write_keyword_list_xlsx(tmp_path):
    path = str(tmp_path / "kw.xlsx")
    write_keyword_list_xlsx(["alpha", "beta", "gamma"], "phrase", "Добавить ключевые слова", path)
    ws = load_workbook(path).active
    flat = [c for row in ws.iter_rows(values_only=True) for c in row if c is not None]
    assert {"alpha", "beta", "gamma"} <= set(flat)
    assert "phrase" in flat  # колонка типа соответствия
