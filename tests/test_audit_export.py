"""Профессиональная выгрузка /audit в Google Sheets (+ .xlsx файлом) — Часть B.

Инварианты (каждый — про честность/деньги, не про красоту):
• 3 вкладки: Находки (весь список worst-first) ПЕРВОЙ + По семьям (свод) + Обзор (карточка)
  ПОСЛЕДНЕЙ — и в Sheets, и в .xlsx (одна раскладка, reports.findings; замечание 5, 2026-07-17:
  файл открывают ради находок, а «Обзор» — дословная копия поста в чате);
• факты находки — отдельными колонками (сортировать/фильтровать), нет факта → пусто, не 0;
• бумага перечисляет ВСЕ быстрые победы (карточка в чате режет блок под кнопки);
• вкладка «По семьям» == result.families (движок посчитал — экспорт не пересчитывает);
• экспорт — БУМАГА: ни колонки suggested_operation, ни кнопки «применить» (золотое правило №3);
• ячейки клиента обезврежены (safe_row: имя кампании `=HYPERLINK(...)` — текст, не формула);
• числовой формат ТОЛЬКО на колонках реестра FINDINGS_FORMATS (иначе проза — как проценты);
• EN-артефакт без кириллицы (RU-утечка), первая вкладка «Findings»;
• publish_audit_to_sheets шарит reader (финансовая бумага), НЕ writer;
• клик по кнопке НЕ пересобирает аудит: холодный кэш → stale-алерт, тёплый → тот же result из кэша.
"""

from __future__ import annotations

import pathlib
import sys
from datetime import date
from types import SimpleNamespace

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

import openpyxl  # noqa: E402
import pytest  # noqa: E402

from audit.engine import build_audit  # noqa: E402
from reports.findings import (  # noqa: E402
    CLICKS_COL,
    COST_COL,
    CPA_COL,
    FAMILY_MONEY_COL,
    FAMILY_SUMMARY_TITLE,
    FINDINGS_TITLE,
    MONEY_COL,
    OVERVIEW_TITLE,
    family_summary_rows,
    findings_meta_rows,
    findings_rows,
)
from reports.period import last_n_days  # noqa: E402
from reports.queries import (  # noqa: E402
    Breakdown,
    GeoWasteRow,
    KeywordQualityRow,
    Metrics,
    SearchTermRow,
)
from reports.sheets import (  # noqa: E402
    SHARE_FAILED,
    build_audit_sheets_data,
    format_requests,
    is_shared,
    publish_audit_to_sheets,
)
from reports.xlsx import build_audit_workbook  # noqa: E402

ROOT = pathlib.Path(__file__).resolve().parents[1]
EVIL = '=HYPERLINK("http://evil","click")'
_CYR = set("абвгдеёжзийклмнопрстуфхцчшщъыьэюяАБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЪЫЬЭЮЯ")


def _report(campaign: str = "Search Brand"):
    """Аккаунт с расходом и БЕЗ конверсий — движку есть что сказать (иначе тест зелен вхолостую)."""
    m = Metrics(impressions=1000, clicks=100, cost_micros=500_000_000, conversions=0.0)
    camp = Breakdown(
        key="campaign",
        title="Кампании",
        dim_headers=["Кампания", "Статус"],
        rows=[((campaign, "ENABLED"), m)],
    )
    return SimpleNamespace(
        customer_id="7753643025",
        totals=m,
        prev_totals=m,
        period=last_n_days(7, today=date(2026, 6, 25)),
        currency="USD",
        breakdowns=[camp],
    )


def _result(campaign: str = "Search Brand"):
    r = build_audit(_report(campaign))
    assert r.findings and r.families, "движку есть что сказать про этот аккаунт"
    return r


def _enriched_result(campaign: str = "Search Brand"):
    """Как gather_audit: тот же result + прицепленное сырьё (report + audit_tables). Данные ЛАТИНИЦЕЙ
    (в EN-артефакте кириллица = наш ярлык, не данные — как в test_export_i18n)."""
    from audit.collect import _geo_waste_table, _keyword_quality_table, _search_terms_table

    r = _result(campaign)
    r.report = _report(campaign)  # ReportData-подобный (breakdowns=[Кампании])
    m = Metrics(impressions=50, clicks=5, cost_micros=10_000_000, conversions=0.0)
    r.audit_tables = [
        t
        for t in (
            _search_terms_table(
                [SearchTermRow("used car export", campaign, "AG", "car", "PHRASE", m)]
            ),
            _keyword_quality_table(
                [
                    KeywordQualityRow(
                        campaign, "AG", "car", 0, "UNSPECIFIED", "AVERAGE", "AVERAGE", m
                    ),
                    KeywordQualityRow(campaign, "AG", "van", 7, "AVERAGE", "AVERAGE", "AVERAGE", m),
                ]
            ),
            _geo_waste_table([GeoWasteRow(campaign, "Tanzania", m)]),
        )
        if t is not None
    ]
    return r


def _cells(rows) -> list[str]:
    return [str(c) for row in rows for c in row if c is not None]


def _xlsx_texts(wb, sheet: str) -> list[str]:
    ws = wb[sheet]
    return [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]


# ── структура: 3 вкладки в Sheets и в .xlsx, «Находки» первой, «Обзор» последней ─────
def test_audit_export_has_three_sections():
    r = _result()
    assert [t.title for t in build_audit_sheets_data(r, "ru")] == [
        FINDINGS_TITLE,
        FAMILY_SUMMARY_TITLE,
        OVERVIEW_TITLE,
    ]
    assert build_audit_workbook(r, "ru").sheetnames == [
        FINDINGS_TITLE,
        FAMILY_SUMMARY_TITLE,
        OVERVIEW_TITLE,
    ]


def test_overview_tab_is_the_audit_card_prose():
    """Вкладка «Обзор» — та же проза, что в карточке /audit (render_audit actions=False), но
    ПОСЛЕДНЕЙ: это копия поста в чате, из-за неё первой выгрузка читалась как «те же данные»."""
    from audit.render import render_audit

    r = _result()
    tab = build_audit_sheets_data(r, "ru")[-1]
    assert tab.title == OVERVIEW_TITLE
    texts = _cells(tab.rows)
    for line in render_audit(r, "ru", actions=False).split("\n"):
        if line:
            assert line in texts, f"обзор карточки не доехал до вкладки: {line}"


# ── вкладка «По семьям» == result.families (движок посчитал, экспорт не пересчитывает) ──
def test_family_summary_matches_result_families():
    from audit.render import family_label

    r = _result()
    rows = family_summary_rows(r, "ru")
    got = {tuple(row) for row in rows}
    exp = set()
    for fam, d in r.families.items():
        at = round(float(d["at_risk"]), 2) if d["at_risk"] > 0 else ""
        exp.add((family_label(fam, "ru"), int(d["count"]), at, round(float(d["penalty"]), 2)))
    assert got == exp, "свод семей разошёлся с движком"
    penalties = [row[3] for row in rows]
    assert penalties == sorted(penalties, reverse=True), "свод должен идти worst-first (по штрафу)"


def test_family_summary_zero_at_risk_is_blank_not_zero():
    """Неденежная семья (at_risk == 0) → пустая ячейка, а НЕ 0.00 (иначе читалось бы как «денег ноль»)."""
    r = _result()
    for row in family_summary_rows(r, "ru"):
        assert row[2] == "" or row[2] > 0  # «Под риском»: либо пусто, либо положительное число


# ── GR3: экспорт — бумага, не кнопка ────────────────────────────────────────────────
def test_audit_export_is_paper_not_a_button():
    """Золотое правило №3: ни в Sheets, ни в .xlsx нет операции находки — только бумага."""
    src = (ROOT / "reports" / "findings.py").read_text(encoding="utf-8")
    assert "suggested_operation" not in src
    r = _result()
    ops = {getattr(f, "suggested_operation", None) for f in r.findings} - {None}
    assert ops, "фикстура без one_tap-находок не проверяет запрет (нужна хоть одна)"
    sheet_cells = set(_cells([row for t in build_audit_sheets_data(r, "ru") for row in t.rows]))
    assert not (sheet_cells & ops), "операция находки утекла во вкладку Sheets"
    for name in build_audit_workbook(r, "ru").sheetnames:
        assert not (set(_xlsx_texts(build_audit_workbook(r, "ru"), name)) & ops)


def test_audit_cells_are_formula_safe(tmp_path):
    """Имя кампании из Google (`=HYPERLINK…`) в ячейке находки → обезврежено safe_row."""
    r = _result(campaign=EVIL)
    path = str(tmp_path / "audit.xlsx")
    build_audit_workbook(r, "ru").save(path)
    wb = openpyxl.load_workbook(path)
    cells = [
        c
        for name in wb.sheetnames
        for row in wb[name].iter_rows()
        for c in row
        if isinstance(c.value, str)
    ]
    for c in cells:
        assert not c.value.startswith(("=", "+", "-", "@")), f"формула в ячейке: {c.value!r}"
    named = [c for c in cells if c.value.lstrip("'") == EVIL]
    assert named, "имя кампании не доехало до книги — тест бесполезен"
    for c in named:
        assert c.data_type == "s" and c.value.startswith("'=")  # текст, НЕ формула


# ── формат: числовые колонки строго по реестру FINDINGS_FORMATS ──────────────────────
def test_money_format_only_on_money_columns():
    r = _result()
    tabs = build_audit_sheets_data(r, "ru")
    reqs = format_requests(tabs)

    def _num_cols(tab_idx: int) -> list[int]:
        return [
            rq["repeatCell"]["range"]["startColumnIndex"]
            for rq in reqs
            if "repeatCell" in rq
            and "numberFormat" in rq["repeatCell"]["cell"]["userEnteredFormat"]
            and rq["repeatCell"]["range"]["sheetId"] == tab_idx
        ]

    # «Находки» — «Под риском» + числовые факты (расход/клики/CPA), и НИЧЕГО сверх реестра
    assert _num_cols(0) == [MONEY_COL, COST_COL, CLICKS_COL, CPA_COL]
    assert _num_cols(1) == [FAMILY_MONEY_COL]  # «По семьям» — только «Под риском»
    assert _num_cols(2) == []  # «Обзор» — проза, числовых колонок нет


# ── колонки фактов листа «Находки» (замечание 5: сортировать/фильтровать, а не парсить прозу) ──
def test_findings_rows_carry_fact_columns():
    """Факты находки — отдельными колонками: cost/clicks/cpa числами, тип соответствия/регион/ключ
    строками; нет факта → ПУСТАЯ ячейка (нет данных ≠ ноль). Ключ и запрос делят одну колонку —
    у находки лежит ровно одно из полей."""
    full = SimpleNamespace(
        severity="warning",
        family="waste",
        target_campaign="Search Brand",
        at_risk=120.5,
        check_id="wasteful_keyword",
        facts={
            "campaign": "Search Brand",
            "cost": 120.456,
            "clicks": 42,
            "cpa": 7.5,
            "match_type": "BROAD",
            "keyword": "buy cars",
        },
    )
    bare = SimpleNamespace(
        severity="info",
        family="structure",
        target_campaign="",
        at_risk=0.0,
        check_id="fake_no_facts",
        facts={},
    )
    row_full, row_bare = findings_rows(SimpleNamespace(currency="USD", findings=[full, bare]), "ru")
    assert row_full[COST_COL] == 120.46
    assert row_full[CLICKS_COL] == 42
    assert row_full[CPA_COL] == 7.5
    assert row_full[CPA_COL + 1] == "BROAD"  # «Тип соответствия»
    assert row_full[CPA_COL + 2] == ""  # «Регион»: факта нет → пусто
    assert row_full[CPA_COL + 3] == "buy cars"  # «Ключ/запрос»
    # у «голой» находки все факт-колонки пустые, НЕ 0 (нет данных ≠ ноль)
    assert [row_bare[c] for c in range(MONEY_COL, CPA_COL + 4)] == [""] * 7
    # поисковый запрос делит колонку «Ключ/запрос» с ключом
    term = SimpleNamespace(
        severity="warning",
        family="search_terms",
        target_campaign="C",
        at_risk=1.0,
        check_id="wasteful_search_term",
        facts={"search_term": "free stuff"},
    )
    (row_term,) = findings_rows(SimpleNamespace(currency="USD", findings=[term]), "ru")
    assert row_term[CPA_COL + 3] == "free stuff"


# ── бумага перечисляет ВСЕ быстрые победы (карточка режет блок под кнопки) ────────────
def test_paper_lists_quick_wins_beyond_the_button_slice():
    """В чате «⚡ Быстрые победы» берутся только из первых QUICK_WIN_POOL находок — ровно им бот
    рисует кнопки, а обещать «в один тап» без кнопки — ложь. У бумаги кнопок нет: обзор выгрузки
    обязан перечислить КАЖДУЮ такую находку (замечание 5, 2026-07-17)."""
    from audit.render import QUICK_WIN_POOL, _quick_win_line, render_audit

    camps = [
        (
            (f"Camp{i:02d}", "ENABLED"),
            Metrics(
                impressions=1000,
                clicks=50,
                cost_micros=(1200 - i * 50) * 1_000_000,
                conversions=0.0,
            ),
        )
        for i in range(12)  # 12 сливающих кампаний > пула кнопок (8)
    ]
    camps.append(
        (
            ("Winner", "ENABLED"),
            Metrics(impressions=1000, clicks=50, cost_micros=100_000_000, conversions=10.0),
        )
    )
    report = SimpleNamespace(
        customer_id="7753643025",
        totals=Metrics(
            impressions=13_000, clicks=650, cost_micros=11_200_000_000, conversions=10.0
        ),
        prev_totals=None,
        period=last_n_days(7, today=date(2026, 6, 25)),
        currency="USD",
        breakdowns=[Breakdown("campaign", "Кампании", ["Кампания", "Статус"], camps)],
    )
    r = build_audit(report)
    one_tap = [f for f in r.findings if f.one_tap]
    beyond = [f for i, f in enumerate(r.findings) if i >= QUICK_WIN_POOL and f.one_tap]
    assert len(one_tap) > QUICK_WIN_POOL and beyond, "фикстура: мало one-tap находок — тест пуст"
    paper = [row[0] for row in findings_meta_rows(r, "ru") if row]
    card = render_audit(r, "ru", actions=False)
    for f in one_tap:  # бумага — инвентарь: строка на КАЖДУЮ
        assert f"• {_quick_win_line(f, 'ru', r.currency)}" in paper
    missed = f"• {_quick_win_line(beyond[-1], 'ru', r.currency)}"
    assert missed in paper and missed not in card  # карточка её не обещает — кнопки нет


# ── EN-артефакт без кириллицы ───────────────────────────────────────────────────────
def test_audit_en_has_no_cyrillic():
    r = _result()
    tabs = build_audit_sheets_data(r, "en")
    assert tabs[0].title == "Findings" and tabs[-1].title == "Overview"
    leaks = [c for c in _cells([row for t in tabs for row in t.rows]) if _CYR & set(c)]
    leaks += [t.title for t in tabs if _CYR & set(t.title)]
    assert not leaks, f"RU-утечка в EN-вкладках Sheets: {leaks}"
    wb = build_audit_workbook(r, "en")
    xleaks = [c for name in wb.sheetnames for c in _xlsx_texts(wb, name) if _CYR & set(c)] + [
        n for n in wb.sheetnames if _CYR & set(n)
    ]
    assert not xleaks, f"RU-утечка в EN-книге: {xleaks}"


# ── обогащённая выгрузка: секция ДАННЫХ (жалоба «в щитс те же данные, что и в посте») ──
def test_converters_shape_and_empty():
    """Конвертеры сырья → Breakdown: форма строк, Metrics берётся как есть, пусто/None → None."""
    from audit.collect import _geo_waste_table, _keyword_quality_table, _search_terms_table

    m = Metrics(impressions=10, clicks=1, cost_micros=1_000_000)

    st = _search_terms_table([SearchTermRow("used car export", "C", "AG", "car", "PHRASE", m)])
    assert st.dim_headers[0] == "Запрос"
    assert st.rows[0][0] == ("used car export", "C", "AG", "car", "PHRASE")
    assert st.rows[0][1] is m  # Metrics не пересобираем — их считает КОД
    assert _search_terms_table(None) is None and _search_terms_table([]) is None

    kq = _keyword_quality_table(
        [
            KeywordQualityRow("C", "AG", "car", 0, "UNSPECIFIED", "AVERAGE", "ABOVE_AVERAGE", m),
            KeywordQualityRow("C", "AG", "van", 7, "AVERAGE", "AVERAGE", "AVERAGE", m),
        ]
    )
    assert kq.rows[0][0][3] == "—"  # QS 0 = нет данных → прочерк, не ложный 0
    assert kq.rows[0][0][4] == ""  # UNSPECIFIED = нет данных → пусто
    assert kq.rows[1][0][3] == 7 and kq.rows[1][0][5] == "AVERAGE"
    assert _keyword_quality_table([]) is None

    geo = _geo_waste_table([GeoWasteRow("C", "Tanzania", m)])
    assert geo.dim_headers == ["Кампания", "Регион"] and geo.rows[0][0] == ("C", "Tanzania")
    assert _geo_waste_table(None) is None


def test_enriched_export_appends_data_section():
    """report прицеплен → между находками/сводом и финальным «Обзором» идут ДАННЫЕ:
    Сводка + разбивки отчёта + запросы/QS/гео."""
    r = _enriched_result()
    titles = [t.title for t in build_audit_sheets_data(r, "ru")]
    assert titles[:2] == [FINDINGS_TITLE, FAMILY_SUMMARY_TITLE]
    assert titles[-1] == OVERVIEW_TITLE
    for expect in ("Сводка", "Кампании", "Поисковые запросы", "Показатель качества", "География"):
        assert expect in titles, f"нет вкладки данных {expect!r}: {titles}"
    names = build_audit_workbook(r, "ru").sheetnames
    assert names[:2] == [FINDINGS_TITLE, FAMILY_SUMMARY_TITLE]
    assert names[-1] == OVERVIEW_TITLE
    for expect in ("Сводка", "Поисковые запросы", "География"):
        assert expect in names, f"нет листа данных {expect!r}: {names}"


def test_no_report_stays_three_tabs():
    """Старый кэш / engine-only вызов (result.report is None) → прежние 3 вкладки, без деградации."""
    r = _result()
    assert getattr(r, "report", None) is None  # build_audit сам не прицепляет
    assert len(build_audit_sheets_data(r, "ru")) == 3
    assert len(build_audit_workbook(r, "ru").sheetnames) == 3


def test_enriched_export_en_has_no_cyrillic():
    """EN-артефакт обогащённой выгрузки без кириллицы (RU-утечка): у новых ярлыков есть EN-пары."""
    r = _enriched_result()
    tabs = build_audit_sheets_data(r, "en")
    leaks = [c for c in _cells([row for t in tabs for row in t.rows]) if _CYR & set(c)]
    leaks += [t.title for t in tabs if _CYR & set(t.title)]
    assert not leaks, f"RU-утечка в EN-вкладках обогащённого экспорта: {leaks}"
    wb = build_audit_workbook(r, "en")
    xleaks = [c for name in wb.sheetnames for c in _xlsx_texts(wb, name) if _CYR & set(c)]
    xleaks += [n for n in wb.sheetnames if _CYR & set(n)]
    assert not xleaks, f"RU-утечка в EN-книге обогащённого экспорта: {xleaks}"


# ── publish_audit_to_sheets: сеть замокана, шарит reader ─────────────────────────────
class _Exec:
    def __init__(self, res):
        self._res = res

    def execute(self):
        return self._res


class _Values:
    def __init__(self, log):
        self.log = log

    def batchUpdate(self, *, spreadsheetId, body):
        self.log.append(("values.batchUpdate", spreadsheetId, body))
        return _Exec({})


class _Spreadsheets:
    def __init__(self, log):
        self.log = log

    def create(self, *, body, fields):
        self.log.append(("create", body, fields))
        return _Exec(
            {
                "spreadsheetId": "SID123",
                "spreadsheetUrl": "https://docs.google.com/spreadsheets/d/SID123",
            }
        )

    def values(self):
        return _Values(self.log)

    def batchUpdate(self, *, spreadsheetId, body):  # форматирование (best-effort)
        self.log.append(("format.batchUpdate", spreadsheetId, body))
        return _Exec({})


class FakeService:
    def __init__(self):
        self.log: list = []

    def spreadsheets(self):
        return _Spreadsheets(self.log)


class _Perms:
    def __init__(self, log, boom=False):
        self.log, self._boom = log, boom

    def create(self, *, fileId, body, fields):
        if self._boom:
            raise RuntimeError("drive down")
        self.log.append(("permissions.create", fileId, body))
        return _Exec({"id": "perm1"})


class FakeDrive:
    def __init__(self, boom=False):
        self.log: list = []
        self._boom = boom

    def permissions(self):
        return _Perms(self.log, self._boom)


def test_publish_audit_to_sheets_shares_reader():
    svc, drive = FakeService(), FakeDrive()
    url, share = publish_audit_to_sheets(_result(), service=svc, drive_service=drive)
    assert url == "https://docs.google.com/spreadsheets/d/SID123"
    assert share == "reader"  # финансовая бумага — reader, НЕ writer
    create = next(e for e in svc.log if e[0] == "create")
    titles = [s["properties"]["title"] for s in create[1]["sheets"]]
    assert titles == [FINDINGS_TITLE, FAMILY_SUMMARY_TITLE, OVERVIEW_TITLE]
    perm = next(e for e in drive.log if e[0] == "permissions.create")
    assert perm[2] == {"type": "anyone", "role": "reader"}


def test_publish_audit_share_failure_degrades_without_raising():
    url, share = publish_audit_to_sheets(
        _result(), service=FakeService(), drive_service=FakeDrive(boom=True)
    )
    assert url.endswith("SID123") and share == SHARE_FAILED and not is_shared(share)


# ── bot-слой: клик не пересобирает аудит (кэш) ──────────────────────────────────────
class _FakeChat:
    id = 4242


class _FakeMessage:
    def __init__(self):
        self.chat = _FakeChat()
        self.answers: list[str] = []

    async def answer(self, text, **kw):
        self.answers.append(text)


async def test_audit_export_cold_cache_is_stale(monkeypatch):
    """Холодный кэш (рестарт бота / старая клавиатура) → stale-алерт; выгрузка НЕ зовётся."""
    import bot.main as bm

    monkeypatch.setattr(bm, "_AUDIT_EXPORT_CACHE", {})
    called: list = []

    async def _no(*a, **kw):
        called.append(a)

    monkeypatch.setattr(bm, "_run_audit_sheets", _no)
    monkeypatch.setattr(bm, "_run_audit_xlsx", _no)
    m = _FakeMessage()
    await bm._run_audit_export(m, "sheets", 4242)
    assert called == []  # публикация не звана (аудит не пересобираем)
    assert m.answers and "audit" in m.answers[0].lower() or "аудит" in m.answers[0].lower()


async def test_audit_export_warm_cache_uses_cached_result(monkeypatch):
    """Тёплый кэш → _run_audit_sheets получает ТОТ ЖЕ result из кэша (пере-сбор аудита не гоняем)."""
    import bot.main as bm

    r = _result()
    monkeypatch.setattr(bm, "_AUDIT_EXPORT_CACHE", {4242: (r, "7753643025")})
    seen: list = []

    async def _sheets(m, result, acct):
        seen.append((result, acct))

    async def _xlsx(m, result, acct):
        seen.append(("xlsx", result, acct))

    monkeypatch.setattr(bm, "_run_audit_sheets", _sheets)
    monkeypatch.setattr(bm, "_run_audit_xlsx", _xlsx)

    await bm._run_audit_export(_FakeMessage(), "sheets", 4242)
    assert seen == [(r, "7753643025")]  # тот же объект — без повторного gather
    assert seen[0][0] is r

    seen.clear()
    await bm._run_audit_export(_FakeMessage(), "xlsx", 4242)
    assert seen == [("xlsx", r, "7753643025")]


if __name__ == "__main__":  # pragma: no cover
    pytest.main([__file__, "-q"])
