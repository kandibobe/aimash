"""XLSX-генератор вложений для больших списков ключей/минус-слов в proposal-сводках.

Когда ключей > confirm.render.KW_INLINE_MAX (20), в сводке пишется «полный список во
вложении .xlsx». Этот модуль создаёт такой файл по требованию — для headless-пути
подтверждения (build_proposal) и для будущих контуров (MCP-WRITE).
"""

from __future__ import annotations

import base64
import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.safe_cell import safe_row  # A9: защита от формула-инъекции

# Стиль шапки: белый жирный шрифт на синем фоне #4472C4 (как в ТЗ).
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_HEADERS = ["Ключевое слово", "Тип соответствия", "Кампания", "Операция"]


def _append(ws, row) -> None:
    """A9: ws.append с обезвреживанием ячеек — защита от CSV/formula injection."""
    ws.append(safe_row(row))


def _autosize(ws, ncols: int, *, cap: int = 60) -> None:
    """Автоширина колонок (как в reports/xlsx.py)."""
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        width = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), cap)


def write_keywords_xlsx(
    keywords: list[str],
    operation: str,
    match_type: str = "",
    campaign: str = "",
) -> str:
    """Сгенерировать .xlsx со списком ключей для proposal-вложения.

    Args:
        keywords: список ключевых слов (или минус-слов).
        operation: код операции (add_keywords, remove_keywords, …).
        match_type: тип соответствия (broad/phrase/exact) — человекочитаемый.
        campaign: имя кампании (опционально).

    Returns:
        Абсолютный путь к файлу .xlsx во временной директории.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None  # новый Workbook всегда имеет активный лист
    ws.title = "Ключевые слова"

    # Заголовки — жирный белый шрифт на синем фоне #4472C4.
    _append(ws, _HEADERS)
    for c in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"

    # Данные построчно.
    for kw in keywords:
        _append(ws, [kw, match_type, campaign, operation])

    _autosize(ws, len(_HEADERS))

    # Сохраняем во временную директорию.
    tmpdir = tempfile.mkdtemp(prefix="aimash_proposal_")
    path = os.path.join(tmpdir, f"proposal_keywords_{operation}.xlsx")
    wb.save(path)
    return path


def read_attachment_b64(path: str) -> str:
    """Прочитать .xlsx-файл и вернуть его содержимое как base64-строку (data URI).

    Используется MCP-слоем для передачи вложения в конверте ответа
    (headless-подтверждение не имеет доступа к файловой системе контейнера).
    После вызова файл и его временную директорию СЛЕДУЕТ удалить через
    cleanup_attachment.
    """
    with open(path, "rb") as fh:
        raw = fh.read()
    return base64.b64encode(raw).decode("ascii")


def cleanup_attachment(path: str) -> None:
    """Удалить .xlsx-файл и его временную директорию (idempotent, best-effort).

    Вызывать после успешного/неуспешного выполнения proposal (finalize/record_failure),
    чтобы не копить временные файлы.
    """
    if not path or not os.path.isfile(path):
        return
    tmpdir = os.path.dirname(path)
    try:
        os.remove(path)
    except OSError:
        pass
    if os.path.isdir(tmpdir):
        try:
            os.rmdir(tmpdir)
        except OSError:
            pass