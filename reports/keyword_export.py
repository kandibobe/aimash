"""Full row-level keyword performance export for Search campaigns."""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ads.client import ensure_read_allowed
from reports.period import Period
from reports.safe_cell import safe_row


@dataclass(frozen=True)
class KeywordPerformanceRow:
    campaign_name: str
    ad_group_name: str
    keyword_text: str
    match_type: str
    impressions: int
    clicks: int
    cost_micros: int
    conversions: float

    @property
    def cost(self) -> float:
        return self.cost_micros / 1_000_000


def _enum_name(value) -> str:
    return getattr(value, "name", str(value))


def keyword_performance_query(period: Period) -> str:
    """GAQL for every keyword row in Search campaigns for one period."""
    return (
        "SELECT campaign.name, ad_group.name, ad_group_criterion.keyword.text, "
        "ad_group_criterion.keyword.match_type, metrics.impressions, metrics.clicks, "
        "metrics.cost_micros, metrics.conversions FROM keyword_view "
        f"WHERE {period.gaql_between()} "
        "AND campaign.advertising_channel_type = 'SEARCH' "
        "ORDER BY campaign.name, ad_group.name, ad_group_criterion.keyword.text"
    )


def fetch_keyword_performance_rows(
    client, customer_id: str, period: Period
) -> list[KeywordPerformanceRow]:
    """Read the complete keyword dataset with SearchStream and no row cap."""
    ensure_read_allowed(customer_id)
    service = client.get_service("GoogleAdsService")
    rows: list[KeywordPerformanceRow] = []
    for batch in service.search_stream(
        customer_id=str(customer_id), query=keyword_performance_query(period)
    ):
        for row in batch.results:
            rows.append(
                KeywordPerformanceRow(
                    campaign_name=str(row.campaign.name),
                    ad_group_name=str(row.ad_group.name),
                    keyword_text=str(row.ad_group_criterion.keyword.text),
                    match_type=_enum_name(row.ad_group_criterion.keyword.match_type),
                    impressions=int(row.metrics.impressions),
                    clicks=int(row.metrics.clicks),
                    cost_micros=int(row.metrics.cost_micros),
                    conversions=float(row.metrics.conversions),
                )
            )
    return rows


_HEADERS = {
    "ru": [
        "Кампания", "Группа объявлений", "Ключевое слово", "Тип соответствия",
        "Показы", "Клики", "Расход", "Конверсии",
    ],
    "en": [
        "Campaign", "Ad group", "Keyword", "Match type",
        "Impressions", "Clicks", "Cost", "Conversions",
    ],
}


def build_keyword_performance_workbook(
    rows: list[KeywordPerformanceRow], *, currency: str = "", language: str = "ru"
) -> Workbook:
    """Create a flat workbook with exactly one data row per keyword criterion."""
    lang = "en" if language == "en" else "ru"
    headers = list(_HEADERS[lang])
    if currency:
        headers[6] = f"{headers[6]}, {currency}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Keywords" if lang == "en" else "Ключевые слова"
    ws.append(headers)
    for item in rows:
        ws.append(
            safe_row(
                [
                    item.campaign_name, item.ad_group_name, item.keyword_text, item.match_type,
                    item.impressions, item.clicks, item.cost, item.conversions,
                ]
            )
        )

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row_number in range(2, ws.max_row + 1):
        ws.cell(row=row_number, column=5).number_format = "#,##0"
        ws.cell(row=row_number, column=6).number_format = "#,##0"
        ws.cell(row=row_number, column=7).number_format = "#,##0.00"
        ws.cell(row=row_number, column=8).number_format = "0.00"
    for column in range(1, len(headers) + 1):
        width = max(
            len(str(ws.cell(row=row, column=column).value or ""))
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 60)
    return wb


def write_keyword_performance_xlsx(
    rows: list[KeywordPerformanceRow], path: str, *, currency: str = "", language: str = "ru"
) -> str:
    build_keyword_performance_workbook(rows, currency=currency, language=language).save(path)
    return path
