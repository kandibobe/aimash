"""Экспорт отчёта (reports.service.ReportData) в .xlsx через openpyxl. READ-ONLY, без секретов.

Лист «Сводка» (итоги + сравнение период-к-периоду) + по листу на каждую разбивку
(кампании/группы/ключи/объявления/устройства/сети/дни). Производные метрики уже посчитаны
кодом (queries.Metrics.as_row) — здесь только раскладка и формат ячеек.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from reports.findings import (
    FAMILY_SUMMARY_FORMATS,
    FAMILY_SUMMARY_TITLE,
    FINDINGS_FORMATS,
    FINDINGS_TITLE,
    MONEY_FORMAT,
    OVERVIEW_TITLE,
    account_health_cells,
    family_summary_headers,
    family_summary_rows,
    findings_headers,
    findings_meta_rows,
    findings_rows,
)
from reports.labels import (
    account_report_title,
    currency_line,
    loc,
    loc_all,
    mcc_deep_title,
    mcc_summary_title,
    period_line,
    top_note,
)
from reports.queries import METRIC_FORMATS, TOP_N, Breakdown, metric_headers
from reports.safe_cell import safe_row  # A9: защита от формула-инъекции в ячейках
from reports.service import ReportData


def _note(b: Breakdown, lang: str) -> str | None:
    """Пометка об усечении на языке пользователя (b.note хранится RU — см. Breakdown.note_kind)."""
    if b.note and b.note_kind:
        return top_note(b.note_kind, TOP_N, lang)
    return b.note


def _append(ws, row) -> None:
    """A9: ws.append с обезвреживанием ячеек (safe_row) — ни один append не пишет сырую
    строку-формулу (=/+/-/@ в имени кампании/ключа/аккаунта из данных клиента)."""
    ws.append(safe_row(row))


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Формат ячеек по позиции в METRIC_HEADERS — реестр один на xlsx и Sheets (reports.queries).
_METRIC_FORMATS = METRIC_FORMATS


def _style_header_row(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"


def _autosize(ws, ncols: int, *, cap: int = 48) -> None:
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        width = max(
            (len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(width + 2, 10), cap)


def _write_breakdown(wb: Workbook, b: Breakdown, currency: str = "", lang: str = "ru") -> None:
    ws = wb.create_sheet(title=loc(b.title, lang)[:31])
    note = _note(b, lang)
    if note:
        _append(ws, [note])  # пометка об усечении — первой строкой
    headers = loc_all(b.dim_headers, lang) + metric_headers(currency, lang)
    header_row = ws.max_row + 1
    _append(ws, headers)
    for dims, m in b.rows:
        _append(ws, list(dims) + m.as_row())
    # Если была строка-пометка, шапка не в первой строке — стилизуем шапку отдельно.
    if header_row == 1:
        _style_header_row(ws, len(headers))
    else:
        for c in range(1, len(headers) + 1):
            ws.cell(row=header_row, column=c).fill = _HEADER_FILL
            ws.cell(row=header_row, column=c).font = _HEADER_FONT
    _apply_metric_formats_at(ws, len(b.dim_headers), header_row + 1, len(b.rows))
    _autosize(ws, len(headers))


def _write_findings(wb: Workbook, result, currency: str = "", lang: str = "ru") -> None:
    """Лист «Находки»: обзор аудита (тот же, что в карточке /audit) + таблица ВСЕХ находок.

    Раскладку даёт reports.findings (общая с Google Sheets), здесь — только запись и формат. Никакой
    колонки «применить»: экспорт — бумага (золотое правило №3)."""
    ws = wb.create_sheet(title=loc(FINDINGS_TITLE, lang)[:31])
    for row in findings_meta_rows(result, lang):
        _append(ws, row)
    _append(ws, [])
    headers = findings_headers(currency, lang)
    _append(ws, headers)
    header_row = ws.max_row
    rows = findings_rows(result, lang)
    for row in rows:
        _append(ws, row)
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).fill = _HEADER_FILL
        ws.cell(row=header_row, column=c).font = _HEADER_FONT
    for col, fmt in FINDINGS_FORMATS:
        for r in range(header_row + 1, header_row + 1 + len(rows)):
            ws.cell(row=r, column=col + 1).number_format = fmt
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _autosize(ws, len(headers), cap=90)  # «Что не так» — предложение, а не ярлык


def _write_simple_table(wb, title, headers, rows, formats, *, cap=48):
    """Лист «шапка (строка 1) + строки данных» с денежным форматом по колонкам formats (0-based).
    Общий раскладчик простых вкладок книги аудита («По семьям», «Находки») — у отчёта шапка часто не
    в первой строке (пометка об усечении), поэтому у него отдельный путь. title — УЖЕ локализован."""
    ws = wb.create_sheet(title=title[:31])
    _append(ws, headers)
    for row in rows:
        _append(ws, row)
    _style_header_row(ws, len(headers))
    for col, fmt in formats:
        for r in range(2, 2 + len(rows)):
            ws.cell(row=r, column=col + 1).number_format = fmt
    _autosize(ws, len(headers), cap=cap)
    return ws


def _apply_metric_formats_at(ws, dim_count: int, first_data_row: int, ndata: int) -> None:
    for i, fmt in enumerate(_METRIC_FORMATS):
        col = dim_count + 1 + i
        for r in range(first_data_row, first_data_row + ndata):
            ws.cell(row=r, column=col).number_format = fmt


def _write_summary(ws, report: ReportData, lang: str = "ru") -> None:
    from reports.period import label_i18n

    p = report.period
    currency = getattr(report, "currency", "") or ""  # defensive: фейк-репорты без поля
    _append(ws, [account_report_title(report.customer_id, lang)])
    _append(ws, [period_line(p, lang)])
    if currency:
        _append(ws, [currency_line(currency, lang)])  # §9: денежные метрики — в валюте аккаунта
    _append(ws, [])
    # Таблица итогов: шапка + строка «текущий период» (+ «предыдущий», если есть сравнение).
    headers = metric_headers(currency, lang)
    _append(ws, [loc("Период", lang), *headers])
    header_row = ws.max_row
    _append(ws, [label_i18n(p, lang), *report.totals.as_row()])
    ndata = 1
    if report.prev_totals is not None:
        _append(ws, [label_i18n(p.previous(), lang), *report.prev_totals.as_row()])
        ndata = 2
    for c in range(1, len(headers) + 2):
        ws.cell(row=header_row, column=c).fill = _HEADER_FILL
        ws.cell(row=header_row, column=c).font = _HEADER_FONT
    _apply_metric_formats_at(ws, 1, header_row + 1, ndata)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _autosize(ws, len(headers) + 1)


def build_workbook(report: ReportData, lang: str = "ru", *, audit=None) -> Workbook:
    """audit — AuditResult (keyword-only ⇒ прежние позиционные вызовы целы). None → книга как раньше,
    без листа «Находки»: собирать аудит — 23 чтения, решает вызыватель (bot._export_audit)."""
    wb = Workbook()
    summary = wb.active
    summary.title = loc("Сводка", lang)
    _write_summary(summary, report, lang)
    currency = getattr(report, "currency", "") or ""  # defensive: фейк-репорты без поля
    if audit is not None:
        _write_findings(wb, audit, currency, lang)  # сразу за «Сводкой»: находки важнее разбивок
    for b in report.breakdowns:
        _write_breakdown(wb, b, currency, lang)
    return wb


def write_report_xlsx(report: ReportData, path: str, lang: str = "ru", *, audit=None) -> str:
    """Сохранить отчёт в .xlsx по пути path. Возвращает path. lang — язык ПОДПИСЕЙ (не данных)."""
    build_workbook(report, lang, audit=audit).save(path)
    return path


def build_audit_workbook(result, lang: str = "ru") -> Workbook:
    """Книга выгрузки /audit из УЖЕ вычисленного AuditResult: «Находки» (весь список worst-first)
    ПЕРВЫМ листом, «По семьям» (свод), секция данных, «Обзор» (карточка без топ-3/дисклеймера,
    проза в колонке A) ПОСЛЕДНИМ — он дословная копия поста в чате (замечание 5, 2026-07-17).
    Зеркало reports.sheets.build_audit_sheets_data. Аудит НЕ пересобираем — на вход готовый result
    (кэш bot-слоя); никаких доп-чтений Google Ads. GR3: колонки «применить» нет — экспорт бумага."""
    wb = Workbook()
    wb.remove(wb.active)  # листы кладём create_sheet'ом в нужном порядке — дефолтный «Sheet» лишний
    currency = getattr(result, "currency", "") or ""
    _write_simple_table(
        wb,
        loc(FINDINGS_TITLE, lang),
        findings_headers(currency, lang),
        findings_rows(result, lang),
        FINDINGS_FORMATS,
        cap=90,  # «Что не так» — предложение, а не ярлык
    )
    _write_simple_table(
        wb,
        loc(FAMILY_SUMMARY_TITLE, lang),
        family_summary_headers(currency, lang),
        family_summary_rows(result, lang),
        FAMILY_SUMMARY_FORMATS,
    )
    # Секция ДАННЫХ (жалоба «в щитс те же данные, что и в посте»): сырьё, прицепленное слоем сбора
    # аудита к result — «Сводка» (итоги) + разбивки отчёта + аудит-таблицы (запросы/QS/гео). Нет
    # report (engine-only вызов / старый кэш) → без секции. Зеркало
    # reports.sheets.build_audit_sheets_data.
    report = getattr(result, "report", None)
    if report is not None:
        rcur = getattr(report, "currency", "") or currency
        _write_summary(wb.create_sheet(title=loc("Сводка", lang)[:31]), report, lang)
        extra = getattr(result, "audit_tables", None) or []
        for b in [*getattr(report, "breakdowns", []), *extra]:
            _write_breakdown(wb, b, rcur, lang)
    overview = wb.create_sheet(title=loc(OVERVIEW_TITLE, lang)[:31])
    for row in findings_meta_rows(result, lang):
        _append(overview, row)
    overview.cell(row=1, column=1).font = Font(bold=True, size=14)
    _autosize(overview, 1, cap=90)  # проза (строки карточки) длиннее ярлыка
    return wb


def write_audit_xlsx(result, path: str, lang: str = "ru") -> str:
    """Сохранить выгрузку /audit в .xlsx по пути path. Возвращает path. lang — язык ПОДПИСЕЙ."""
    build_audit_workbook(result, lang).save(path)
    return path


# ── MCC-сводка (§8): подытоги по валюте + лист аккаунтов + лист пропусков/ошибок ────
def _mcc_campaign_status_cell(cr) -> str:
    """§8: строка разбивки кампаний по статусу для xlsx («ENABLED:3, PAUSED:5»). Fallback на
    ENABLED-счётчик (active_campaigns), если полной разбивки нет; '' если не прочитано."""
    sc = getattr(cr, "campaign_status", None)
    if isinstance(sc, dict) and sc:
        return ", ".join(f"{k}:{v}" for k, v in sorted(sc.items()))
    n = getattr(cr, "active_campaigns", None)
    return f"ENABLED:{n}" if n is not None else ""


def build_mcc_workbook(summary, lang: str = "ru") -> Workbook:
    """Книга по сводке дочерних MCC (reports.mcc.MccSummary, duck-typed). Листы: «Сводка MCC»
    (подытоги ПО ВАЛЮТАМ — без FX), «Аккаунты» (по строке на лист-аккаунт), «Пропущено/ошибки»
    (read-list/менеджерские/частичные сбои — без тихого замалчивания, §8/§5)."""
    wb = Workbook()
    p = summary.period

    # 1) Сводка MCC: подытоги по валюте (валюта — отдельной колонкой, т.к. денежные колонки
    # разных валют нельзя свести в один заголовок; FX не делаем).
    ws = wb.active
    ws.title = loc("Сводка MCC", lang)
    _append(ws, [mcc_summary_title(summary.manager_id, lang)])
    _append(ws, [period_line(p, lang)])
    _append(ws, [])
    headers = [loc("Валюта", lang), loc("Аккаунтов", lang), *metric_headers("", lang)]
    _append(ws, headers)
    header_row = ws.max_row
    for sub in summary.subtotals:
        _append(ws, [sub.currency, sub.accounts, *sub.totals.as_row()])
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).fill = _HEADER_FILL
        ws.cell(row=header_row, column=c).font = _HEADER_FONT
    _apply_metric_formats_at(ws, 2, header_row + 1, len(summary.subtotals))
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _autosize(ws, len(headers))

    # 2) Аккаунты: по строке на лист-аккаунт (id/имя/валюта/статус + метрики).
    acc = wb.create_sheet(title=loc("Аккаунты", lang))
    # §8: колонка «Кампании (статус)» — разбивка по статусу (ENABLED:n, PAUSED:m) перед метриками.
    acc_headers = [
        *loc_all(["ID", "Аккаунт", "Валюта", "Статус", "Кампании (статус)"], lang),
        *metric_headers("", lang),
    ]
    _append(acc, acc_headers)
    for cr in summary.children:
        a = cr.account
        _append(
            acc,
            [
                a.id,
                getattr(a, "name", "") or "",
                a.currency or "",
                getattr(a, "status", "") or "",
                _mcc_campaign_status_cell(cr),
                *cr.totals.as_row(),
            ],
        )
    _style_header_row(acc, len(acc_headers))
    _apply_metric_formats_at(acc, 5, 2, len(summary.children))
    _autosize(acc, len(acc_headers))

    # 3) Пропущено/ошибки: read-list пропуски, менеджерские, частичные сбои чтения.
    _write_issues(wb, summary, lang)
    return wb


def _write_issues(wb: Workbook, src, lang: str) -> None:
    """Лист «Пропущено и ошибки» — общий для MCC-книги и DEEP-книги (§8/§5: без замалчивания)."""
    iss = wb.create_sheet(title=loc("Пропущено и ошибки", lang))
    _append(iss, loc_all(["Тип", "Аккаунт", "Причина"], lang))
    _style_header_row(iss, 3)
    for ch in getattr(src, "inactive", []):
        name = getattr(ch, "name", "") or ch.id
        _append(
            iss,
            [
                loc("неактивный (не ENABLED)", lang),
                f"{name} ({ch.id})",
                getattr(ch, "status", "") or "",
            ],
        )
    for cid in src.skipped:
        _append(iss, [loc("пропущен (нет доступа на чтение)", lang), cid, ""])
    for cid in src.managers:
        _append(iss, [loc("менеджерский (без собственных метрик)", lang), cid, ""])
    for cid, reason in src.errors:
        _append(iss, [loc("ошибка чтения", lang), cid, reason])
    _autosize(iss, 3)


def write_mcc_xlsx(summary, path: str, lang: str = "ru") -> str:
    """Сохранить MCC-сводку в .xlsx по пути path. Возвращает path."""
    build_mcc_workbook(summary, lang).save(path)
    return path


# ── 2.2 (аудит 2026-07-06): DEEP-книга по всем дочерним MCC — лист на аккаунт ─────────
_SHEET_BAD_CHARS = set("\\/*?:[]")


def _sheet_title(name: str, cid: str) -> str:
    """Имя листа openpyxl: «Имя_<customer_id>», санитизация запрещённых символов, ≤31 симв.

    ПОЛНЫЙ id (а не last4): id уникален ⇒ два аккаунта не могут дать одно имя листа. Раньше
    «имя[:24]_last4» совпадали у аккаунтов с похожими именами, а дедуп-цикл при этом зависал.
    """
    tail = "".join(ch for ch in str(cid or "") if ch.isdigit())[:20] or "acct"
    base = (name or "").strip()
    base = "".join(("_" if ch in _SHEET_BAD_CHARS else ch) for ch in base)
    head = base[: 30 - len(tail)]
    return f"{head}_{tail}"[:31] if head else tail[:31]


def _unique_sheet_title(title: str, used: set[str]) -> str:
    """Уникализация имени листа с ОГРАНИЧЕННЫМ счётчиком.

    Раньше стояло `while title in used: title = (title[:29] + "_x")[:31]` — для 31-символьного
    title это тождество (фиксированная точка) ⇒ бесконечный цикл в потоке пула (вызов шёл без
    таймаута → поток занят навсегда). Сейчас коллизия невозможна по построению (полный id),
    но уникализатор оставляем как гард — с гарантией завершения.
    """
    if title not in used:
        return title
    for i in range(2, 100):
        suffix = f"~{i}"
        cand = f"{title[: 31 - len(suffix)]}{suffix}"
        if cand not in used:
            return cand
    raise ValueError(f"не удалось подобрать уникальное имя листа для {title!r}")


def build_mcc_deep_workbook(deep, lang: str = "ru") -> Workbook:
    """Книга DEEP-отчёта (reports.mcc.MccDeep, duck-typed): лист «Сводка» (строка на аккаунт,
    метрики в валюте аккаунта — БЕЗ FX) → по ЛИСТУ НА АККАУНТ (итоги+сравнение периода, как
    write_report_xlsx, + таблица разбивки по кампаниям — не 8 листов × N акк.) → «Пропущено и
    ошибки» (без тихого замалчивания)."""
    wb = Workbook()
    p = deep.period

    ws = wb.active
    ws.title = loc("Сводка", lang)
    _append(ws, [mcc_deep_title(deep.manager_id, lang)])
    _append(ws, [period_line(p, lang)])
    _append(ws, [])
    # Здоровье (балл/грейд/под риском) — движком по УЖЕ собранным отчётам: 0 доп. чтений на N
    # аккаунтов. Полный аудит веером (23 чтения × N) сюда НЕ заводим — квота (см. reports.findings).
    headers = [
        *loc_all(["ID", "Аккаунт", "Валюта"], lang),
        *metric_headers("", lang),
        *loc_all(["Балл", "Оценка", "Под риском"], lang),
    ]
    _append(ws, headers)
    header_row = ws.max_row
    for ch, report in deep.items:
        _append(
            ws,
            [
                ch.id,
                getattr(ch, "name", "") or "",
                ch.currency or "",
                *report.totals.as_row(),
                *account_health_cells(report),
            ],
        )
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).fill = _HEADER_FILL
        ws.cell(row=header_row, column=c).font = _HEADER_FONT
    _apply_metric_formats_at(ws, 3, header_row + 1, len(deep.items))
    for r in range(header_row + 1, header_row + 1 + len(deep.items)):
        ws.cell(row=r, column=len(headers)).number_format = MONEY_FORMAT  # «Под риском»
    _append(ws, [])
    _append(ws, [loc("Балл — по данным отчёта (без полного аудита); полный разбор — /audit", lang)])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _autosize(ws, len(headers))

    used_titles = {ws.title}
    for ch, report in deep.items:
        title = _unique_sheet_title(_sheet_title(getattr(ch, "name", "") or "", ch.id), used_titles)
        used_titles.add(title)
        acc_ws = wb.create_sheet(title=title)
        _write_summary(acc_ws, report, lang)
        camp = next((b for b in report.breakdowns if b.key == "campaign"), None)
        if camp and camp.rows:
            _append(acc_ws, [])
            camp_headers = loc_all(camp.dim_headers, lang) + metric_headers(
                report.currency or "", lang
            )
            _append(acc_ws, camp_headers)
            hrow = acc_ws.max_row
            for dims, m in camp.rows:
                _append(acc_ws, list(dims) + m.as_row())
            for c in range(1, len(camp_headers) + 1):
                acc_ws.cell(row=hrow, column=c).fill = _HEADER_FILL
                acc_ws.cell(row=hrow, column=c).font = _HEADER_FONT
            _apply_metric_formats_at(acc_ws, len(camp.dim_headers), hrow + 1, len(camp.rows))
            _autosize(acc_ws, len(camp_headers))

    _write_issues(wb, deep, lang)
    return wb


def write_mcc_deep_xlsx(deep, path: str, lang: str = "ru") -> str:
    """Сохранить DEEP-книгу по пути path. Возвращает path."""
    build_mcc_deep_workbook(deep, lang).save(path)
    return path


def build_budget_projection_workbook(spec, lang: str = "ru") -> Workbook:
    """График L3 к карточке подтверждения (Волна 5): накопленный расход при прежнем и новом дневном
    бюджете. На вход — `confirm.attachment.BudgetChartSpec` (duck-typed, чтобы `reports/` не зависел
    от `confirm/`).

    Две ПРЯМЫЕ линии, и лист говорит об этом первой же строкой. Модели поведения кампании в аукционе
    у нас нет; рисовать кривую «прогноза» означало бы выдать арифметику дневного потолка за
    предсказание — ровно тот класс, который правило 4/15 запрещает на денежном пути. Расхождение
    линий на последнем дне равно `delta_horizon_micros` из текста карточки: график и текст считаются
    из одних чисел и разойтись не могут.

    Ни одного чтения Google Ads: курьер вложений (`scheduler.jobs.deliver_proposal_attachments`)
    SDK не трогает по контракту, файл собирается из `params` той же строки черновика."""
    from confirm.consequences import PROJECTION_DAYS, projection_rows

    cons = spec.cons
    en = lang == "en"
    wb = Workbook()
    ws = wb.active
    ws.title = "Spend projection" if en else "Проекция расхода"
    head = (
        f"Campaign “{spec.campaign}” — daily budget projection"
        if en
        else f"Кампания «{spec.campaign}» — проекция расхода по дневному бюджету"
    )
    _append(ws, [head])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    _append(
        ws,
        [
            "Linear projection of the daily cap, not an auction forecast."
            if en
            else "Линейная проекция дневного потолка, не прогноз аукциона."
        ],
    )
    if cons.shared_campaigns:
        # Общий бюджет: линия проекции описывает не одну кампанию, а весь набор на этом бюджете.
        # Без этой строки лист отвечал бы на вопрос про кампанию, а рисовал бы про чужие деньги.
        n = len(cons.shared_campaigns)
        _append(
            ws,
            [
                f"Shared budget: the projection covers {n} more campaign(s) on it."
                if en
                else f"Общий бюджет: проекция описывает ещё {n} кампан. на нём."
            ],
        )
    _append(ws, [])

    cur = cons.currency or ""
    headers = (
        [
            "Day",
            f"At previous budget{f' ({cur})' if cur else ''}",
            f"At new budget{f' ({cur})' if cur else ''}",
        ]
        if en
        else [
            "День",
            f"При прежнем бюджете{f' ({cur})' if cur else ''}",
            f"При новом бюджете{f' ({cur})' if cur else ''}",
        ]
    )
    _append(ws, headers)
    header_row = ws.max_row
    for c in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=c).fill = _HEADER_FILL
        ws.cell(row=header_row, column=c).font = _HEADER_FONT
    for day, before_micros, after_micros in projection_rows(cons, PROJECTION_DAYS):
        _append(ws, [day, before_micros / 1_000_000, after_micros / 1_000_000])
    last_row = ws.max_row
    for r in range(header_row + 1, last_row + 1):
        for c in (2, 3):
            ws.cell(row=r, column=c).number_format = MONEY_FORMAT

    chart = LineChart()
    chart.title = "Cumulative spend" if en else "Накопленный расход"
    chart.y_axis.title = cur or ("money" if en else "деньги")
    chart.x_axis.title = "Day" if en else "День"
    chart.height, chart.width = 9, 18
    # titles_from_data=True — подписи серий берутся из строки заголовков, поэтому диапазон данных
    # начинается ИМЕННО с неё, а категории — со следующей.
    chart.add_data(
        Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=last_row),
        titles_from_data=True,
    )
    chart.set_categories(Reference(ws, min_col=1, min_row=header_row + 1, max_row=last_row))
    ws.add_chart(chart, "F4")
    _autosize(ws, len(headers))
    return wb


def write_budget_projection_xlsx(spec, path: str, lang: str = "ru") -> str:
    """Сохранить книгу проекции бюджета по пути path. Возвращает path."""
    build_budget_projection_workbook(spec, lang).save(path)
    return path
